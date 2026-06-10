from flask import send_from_directory, Flask, jsonify, request
from flask_cors import CORS
import yfinance as yf
import requests
import feedparser
import anthropic
import json
import os
from dotenv import load_dotenv
load_dotenv()
import time
import re
import threading
import zipfile
import io
import xml.etree.ElementTree as ET
from datetime import date, timedelta, datetime
import pandas as pd
import telegram_report
import dart_report
import earnings_tracker

app = Flask(__name__)

if os.environ.get("TG_API_ID") and os.environ.get("TG_API_HASH"):
    telegram_report.start_realtime_watcher(
        os.environ.get("TG_API_ID"),
        os.environ.get("TG_API_HASH"),
        os.environ.get("TG_SESSION_PATH", "sessions/tg_report"),
    )
    telegram_report.start_index_scheduler(
        os.environ.get("TG_API_ID"),
        os.environ.get("TG_API_HASH"),
        os.environ.get("TG_SESSION_PATH", "sessions/tg_report"),
    )
CORS(app, resources={r"/api/*": {"origins": "*"}})

# 티커 매핑
MAIN_TICKERS = {
    "sp500": "^GSPC",
    "nasdaq": "^IXIC",
    "vix": "^VIX",
    "usdkrw": "KRW=X",
    "wti": "CL=F",
    "gold": "GC=F",
}

EXTRA_TICKERS = {
    "us10y": "^TNX",
    "us2y": "^IRX",
    "dxy": "DX-Y.NYB",
    "brent": "BZ=F",
    "bitcoin": "BTC-USD",
}

SECTOR_TICKERS = {
    "에너지": "XLE",
    "헬스케어": "XLV",
    "부동산": "XLRE",
    "필수소비재": "XLP",
    "자유소비재": "XLY",
    "금융": "XLF",
    "통신서비스": "XLC",
    "산업재": "XLI",
    "정보기술": "XLK",
    "소재": "XLB",
    "유틸리티": "XLU",
    "M7": "MAGS",
    "반도체": "SMH",
}


def fetch_batch_data(ticker_map):
    """여러 티커를 한 번에 다운로드하여 현재가/등락률을 반환한다."""
    symbols = list(ticker_map.values())
    try:
        data = yf.download(symbols, period="5d", progress=False, threads=False)
        if data.empty:
            return {k: {"price": 0, "change": 0, "change_pct": 0} for k in ticker_map}
    except Exception:
        return {k: {"price": 0, "change": 0, "change_pct": 0} for k in ticker_map}

    result = {}
    for key, symbol in ticker_map.items():
        try:
            close_col = data["Close"]
            if isinstance(close_col, pd.DataFrame):
                closes = close_col[symbol].dropna() if symbol in close_col.columns else close_col.iloc[:, 0].dropna()
            else:
                closes = close_col.dropna()
            if len(closes) < 2:
                result[key] = {"price": 0, "change": 0, "change_pct": 0}
                continue
            current = closes.iloc[-1]
            previous = closes.iloc[-2]
            change = current - previous
            change_pct = (change / previous) * 100
            result[key] = {
                "price": round(float(current), 2),
                "change": round(float(change), 2),
                "change_pct": round(float(change_pct), 2),
            }
        except Exception:
            result[key] = {"price": 0, "change": 0, "change_pct": 0}
    return result


@app.route("/", methods=["GET"])
def index():
    return send_from_directory(".", "index.html")

@app.route("/api/market", methods=["GET"])
def get_market_data():
    """핵심 지표 6개 (S&P500, NASDAQ, VIX, USD/KRW, WTI, Gold)"""
    return jsonify(fetch_batch_data(MAIN_TICKERS))


@app.route("/api/extra", methods=["GET"])
def get_extra_data():
    """채권/환율/원자재 추가 데이터"""
    return jsonify(fetch_batch_data(EXTRA_TICKERS))


@app.route("/api/sectors", methods=["GET"])
def get_sector_data():
    """섹터별 등락률"""
    sector_data = fetch_batch_data(SECTOR_TICKERS)
    sectors = []
    for name in SECTOR_TICKERS:
        d = sector_data.get(name, {})
        sectors.append({"name": name, "value": d.get("change_pct", 0)})
    sectors.sort(key=lambda x: x["value"], reverse=True)
    return jsonify(sectors)


@app.route("/api/fear-greed", methods=["GET"])
def get_fear_greed():
    """CNN Fear & Greed Index 실제 데이터"""
    try:
        res = requests.get(
            "https://production.dataviz.cnn.io/index/fearandgreed/graphdata",
            headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
                "Accept": "application/json",
                "Referer": "https://edition.cnn.com/markets/fear-and-greed",
                "Origin": "https://edition.cnn.com",
            },
            timeout=10,
        )
        data = res.json()
        score = round(data["fear_and_greed"]["score"])
        rating = data["fear_and_greed"]["rating"]
        rating_ko = {
            "Extreme Fear": "극단적 공포",
            "Fear": "공포",
            "Neutral": "중립",
            "Greed": "탐욕",
            "Extreme Greed": "극단적 탐욕",
        }.get(rating, rating)
        return jsonify({"value": score, "rating": rating_ko})
    except Exception:
        # 폴백: VIX 기반 간이 계산
        try:
            vix_data = fetch_batch_data({"vix": "^VIX"})
            vix = vix_data.get("vix", {}).get("price", 0)
            if vix > 0:
                score = max(0, min(100, int(100 - ((vix - 10) / 30) * 100)))
                return jsonify({"value": score, "rating": "VIX 기반 추정"})
        except Exception:
            pass
        return jsonify({"value": 50, "rating": "데이터 없음"})


_stock_detail_cache = {}


def _find_in_market_stocks(ticker):
    """MARKET_STOCKS에서 ticker에 해당하는 (name, sector, industry)를 찾는다."""
    for market, stocks in MARKET_STOCKS.items():
        for t, name, sector, industry in stocks:
            if t == ticker:
                return name, sector, industry
    return None, None, None


@app.route("/api/stock-detail/<ticker>", methods=["GET"])
def get_stock_detail(ticker):
    """종목 클릭 시 기업 개요 + 등락 이유 뉴스 요약."""
    today_key = date.today().isoformat()
    cache_key = (ticker, today_key)

    stale = [k for k in _stock_detail_cache if k[1] != today_key]
    for k in stale:
        del _stock_detail_cache[k]

    if cache_key in _stock_detail_cache:
        return jsonify(_stock_detail_cache[cache_key])

    change_pct = request.args.get("change_pct", "")
    is_korean = ticker.endswith(".KS") or ticker.endswith(".KQ")
    ms_name, ms_sector, ms_industry = _find_in_market_stocks(ticker)

    # --- 기업 정보 조회 ---
    yf_name, yf_summary, yf_sector, yf_industry = "", "", "", ""
    try:
        info = yf.Ticker(ticker).info
        yf_name = info.get("longName") or info.get("shortName") or ""
        yf_summary = info.get("longBusinessSummary") or ""
        yf_sector = info.get("sector") or ""
        yf_industry = info.get("industry") or ""
    except Exception:
        pass

    display_name = ms_name or yf_name or ticker
    sector = ms_sector or yf_sector
    industry = ms_industry or yf_industry

    SECTOR_KO = {
        "Technology": "정보기술", "Information Technology": "정보기술",
        "Healthcare": "헬스케어", "Health Care": "헬스케어",
        "Financials": "금융", "Financial Services": "금융",
        "Consumer Cyclical": "경기소비재", "Consumer Discretionary": "경기소비재",
        "Consumer Defensive": "필수소비재", "Consumer Staples": "필수소비재",
        "Energy": "에너지", "Industrials": "산업재",
        "Communication Services": "통신서비스", "Materials": "소재",
        "Real Estate": "부동산", "Utilities": "유틸리티",
    }
    if sector and not any(ord(c) > 127 for c in sector):
        sector = SECTOR_KO.get(sector, sector)

    overview = ""
    try:
        if yf_summary:
            resp = claude_client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=300,
                messages=[{"role": "user", "content": (
                    f"다음 영문 기업 소개를 한국어 2-3문장으로 요약해라. "
                    f"회사명: {display_name}\n\n{yf_summary[:1500]}\n\n"
                    f"답변에 마크다운 문법(#, *, **, -, 등)이나 제목/헤더를 절대 포함하지 마. 평문 한국어로만 답변해."
                )}],
            )
            overview = resp.content[0].text.strip()
        else:
            ctx = f"회사명: {display_name}"
            if sector:
                ctx += f", 섹터: {sector}"
            if industry:
                ctx += f", 업종: {industry}"
            resp = claude_client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=300,
                messages=[{"role": "user", "content": (
                    f"{ctx}\n이 종목의 사업 개요를 한국어 2-3문장으로 작성해라.\n\n"
                    f"답변에 마크다운 문법(#, *, **, -, 등)이나 제목/헤더를 절대 포함하지 마. 평문 한국어로만 답변해."
                )}],
            )
            overview = resp.content[0].text.strip()
    except Exception:
        overview = "기업 정보를 불러올 수 없습니다"

    overview = overview.strip()
    lines = overview.split('\n')
    lines = [l for l in lines if not l.strip().startswith('#')]
    overview = '\n'.join(lines).strip()

    # --- 뉴스 검색 ---
    articles = []
    try:
        if is_korean:
            ko_name = ms_name or yf_name or ticker
            articles = fetch_stock_news(f"{ko_name} 주가", lang="ko")
        else:
            en_name = yf_name or ms_name or ticker
            articles = fetch_stock_news(f"{en_name} stock", lang="en")
    except Exception:
        articles = []

    # --- 뉴스 기반 등락 이유 요약 ---
    news_summary = []
    news_sources = []
    if articles:
        news_sources = [
            {"title": a["title"], "url": a["url"], "source": a["source"]}
            for a in articles[:3]
        ]
        article_text = "\n".join(
            f"- [{a['source']}] {a['title']}" for a in articles[:5]
        )
        change_ctx = f"오늘 이 종목이 {change_pct}% 움직였는데, " if change_pct else ""
        try:
            resp = claude_client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=400,
                messages=[{"role": "user", "content": (
                    f"다음은 {display_name} 관련 최근 뉴스다.\n{article_text}\n\n"
                    f"{change_ctx}이 움직임의 이유로 가장 가능성이 높은 것을 한국어로 요약해라. "
                    f"뉴스에 명확한 이유가 없으면 '- 특별한 뉴스 동인이 확인되지 않음'으로 답변.\n\n"
                    f"답변은 반드시 다음 형식을 정확히 따라:\n"
                    f"- 첫 번째 이유 한 문장\n"
                    f"- 두 번째 이유 한 문장\n"
                    f"- 세 번째 이유 한 문장 (선택)\n\n"
                    f"각 줄은 정확히 '- '로 시작하고, 그 외에 #, *, **, 제목 같은 마크다운 문법이나 추가 설명은 절대 포함하지 마."
                )}],
            )
            raw = resp.content[0].text.strip()
            news_summary = [line.strip() for line in raw.split("\n")
                           if line.strip() and not line.strip().startswith('#')]
        except Exception:
            news_summary = ["- 뉴스 요약을 생성할 수 없습니다"]
    else:
        news_summary = ["- 관련 뉴스를 찾을 수 없습니다"]

    result = {
        "ticker": ticker,
        "name": display_name,
        "sector": sector,
        "industry": industry,
        "overview": overview,
        "news_summary": news_summary,
        "news_sources": news_sources,
    }
    _stock_detail_cache[cache_key] = result
    return jsonify(result)


@app.route("/api/stock-opinion/<ticker>", methods=["GET"])
def get_stock_opinion(ticker):
    is_korean = ticker.endswith(".KS") or ticker.endswith(".KQ")
    if not is_korean:
        return jsonify({"available": False})

    try:
        from kis_api import get_invest_opinion
        df = get_invest_opinion(ticker, months=6)
    except Exception:
        return jsonify({"available": False})

    if df is None or (hasattr(df, 'empty') and df.empty):
        return jsonify({"available": False})

    df_sorted = df.reset_index().sort_values('date', ascending=False)
    current_price = int(df_sorted.iloc[0]['close_price']) if len(df_sorted) > 0 else 0
    latest = df_sorted.groupby('broker').first()

    _OPINION_MAP = {
        'buy': 'buy', '매수': 'buy', 'strong buy': 'buy',
        'hold': 'hold', '중립': 'hold', '보유': 'hold', 'neutral': 'hold',
        'marketperform': 'hold', 'market perform': 'hold',
        'sell': 'sell', '매도': 'sell', 'underperform': 'sell',
        'strong sell': 'sell', 'reduce': 'sell',
    }

    buy_count = 0
    hold_count = 0
    sell_count = 0
    for op in latest['opinion']:
        mapped = _OPINION_MAP.get(str(op).strip().lower(), 'buy')
        if mapped == 'buy':
            buy_count += 1
        elif mapped == 'hold':
            hold_count += 1
        else:
            sell_count += 1

    targets = latest['target_price']
    targets = targets[targets > 0]

    broker_opinions = []
    for broker_name, row in latest.iterrows():
        broker_opinions.append({
            "broker": broker_name,
            "opinion": row["opinion"],
            "target_price": int(row["target_price"]),
            "date": row["date"].strftime("%Y-%m-%d") if hasattr(row["date"], "strftime") else str(row["date"])[:10],
        })
    broker_opinions.sort(key=lambda x: x["date"], reverse=True)

    result = {
        "available": True,
        "ticker": ticker,
        "target_price_avg": int(targets.mean()) if len(targets) > 0 else 0,
        "target_price_high": int(targets.max()) if len(targets) > 0 else 0,
        "target_price_low": int(targets.min()) if len(targets) > 0 else 0,
        "current_price": current_price,
        "consensus": {"buy": buy_count, "hold": hold_count, "sell": sell_count},
        "coverage_count": len(latest),
        "broker_opinions": broker_opinions,
    }
    return jsonify(result)


@app.route("/api/stock-moat/<ticker>", methods=["GET"])
def get_stock_moat(ticker):
    is_korean = ticker.endswith(".KS") or ticker.endswith(".KQ")
    if not is_korean:
        return jsonify({"has_data": False})
    ms_name, _, _ = _find_in_market_stocks(ticker)
    if not ms_name:
        try:
            info = yf.Ticker(ticker).info
            ms_name = info.get("longName") or info.get("shortName") or ticker
        except Exception:
            ms_name = ticker
    try:
        from moat_analyzer import get_or_analyze_moat
        result = get_or_analyze_moat(ticker, ms_name)
    except Exception:
        return jsonify({"has_data": False})
    if not result:
        return jsonify({"has_data": False})
    return jsonify({
        "has_data": True,
        "ticker": ticker,
        "name": result.get("name", ms_name),
        "moat": result.get("moat", ""),
        "bottleneck": result.get("bottleneck", ""),
        "news_count": result.get("news_count", 0),
        "last_updated": result.get("last_updated", ""),
    })


@app.route("/api/stock-quarterly-eps/<ticker>", methods=["GET"])
def get_stock_quarterly_eps(ticker):
    is_korean = ticker.endswith(".KS") or ticker.endswith(".KQ")
    if not is_korean:
        return jsonify({"has_data": False})
    stock_code = ticker.replace(".KS", "").replace(".KQ", "")
    cache_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cache", "stock_quarterly_eps.json")
    if not os.path.exists(cache_path):
        return jsonify({"has_data": False})
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            cache = json.load(f)
    except Exception:
        return jsonify({"has_data": False})
    entry = cache.get(stock_code)
    if not entry or not entry.get("eps_series"):
        return jsonify({"has_data": False})
    return jsonify({
        "has_data": True,
        "ticker": ticker,
        "series": entry["eps_series"],
    })


@app.route("/api/chart", methods=["GET"])
def get_chart():
    """차트용 히스토리컬 데이터"""
    from datetime import datetime, timedelta
    symbol = request.args.get("symbol", "^GSPC")
    period = request.args.get("period", "1y")
    start = request.args.get("start", "")
    end = request.args.get("end", "")
    use_range = False
    dl_start = None
    dl_end = None
    if start and end:
        try:
            dt_s = datetime.strptime(start, "%Y-%m-%d")
            dt_e = datetime.strptime(end, "%Y-%m-%d")
            if dt_s < dt_e:
                dl_start = dt_s.strftime("%Y-%m-%d")
                dl_end = (dt_e + timedelta(days=1)).strftime("%Y-%m-%d")
                use_range = True
        except ValueError:
            use_range = False
    if not use_range and period not in {"5y", "2y", "1y", "6mo", "ytd"}:
        period = "1y"
    try:
        if use_range:
            data = yf.download(symbol, start=dl_start, end=dl_end, interval="1d", progress=False, threads=False)
        else:
            data = yf.download(symbol, period=period, progress=False, threads=False)
        if data.empty:
            return jsonify({"dates": [], "prices": []})
        # yfinance MultiIndex 컬럼 처리
        if isinstance(data.columns, pd.MultiIndex):
            data.columns = data.columns.get_level_values(0)
        dates = data.index.strftime("%Y-%m-%d").tolist()
        prices = [round(float(p), 2) for p in data["Close"]]
        return jsonify({"dates": dates, "prices": prices})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"dates": [], "prices": []})


# 시장별 주요 종목 (ticker, name, sector, industry)
MARKET_STOCKS = {
    "us": [
        ("AAPL", "Apple", "정보기술", "소비자전자"), ("MSFT", "Microsoft", "정보기술", "소프트웨어"),
        ("GOOGL", "Alphabet", "통신서비스", "인터넷"), ("AMZN", "Amazon", "경기소비재", "이커머스"),
        ("NVDA", "Nvidia", "정보기술", "반도체"), ("META", "Meta", "통신서비스", "소셜미디어"),
        ("TSLA", "Tesla", "경기소비재", "전기차"), ("JPM", "JP Morgan", "금융", "은행"),
        ("V", "Visa", "금융", "결제"), ("UNH", "UnitedHealth", "헬스케어", "보험"),
        ("XOM", "ExxonMobil", "에너지", "석유"), ("JNJ", "J&J", "헬스케어", "제약"),
        ("WMT", "Walmart", "필수소비재", "유통"), ("PG", "P&G", "필수소비재", "생활용품"),
        ("MA", "Mastercard", "금융", "결제"), ("HD", "Home Depot", "경기소비재", "주택개선"),
        ("CVX", "Chevron", "에너지", "석유"), ("MRK", "Merck", "헬스케어", "제약"),
        ("ABBV", "AbbVie", "헬스케어", "바이오"), ("LLY", "Eli Lilly", "헬스케어", "제약"),
        ("AVGO", "Broadcom", "정보기술", "반도체"), ("PEP", "PepsiCo", "필수소비재", "음료"),
        ("KO", "Coca-Cola", "필수소비재", "음료"), ("MCD", "McDonald's", "경기소비재", "외식"),
        ("CRM", "Salesforce", "정보기술", "소프트웨어"), ("AMD", "AMD", "정보기술", "반도체"),
        ("BA", "Boeing", "산업재", "항공우주"), ("GS", "Goldman Sachs", "금융", "투자은행"),
        ("CAT", "Caterpillar", "산업재", "중장비"), ("NEE", "NextEra", "유틸리티", "전력"),
        # 나스닥 주요 종목
        ("MU", "Micron", "정보기술", "반도체"), ("INTC", "Intel", "정보기술", "반도체"),
        ("QCOM", "Qualcomm", "정보기술", "반도체"), ("TXN", "Texas Instruments", "정보기술", "반도체"),
        ("MRVL", "Marvell", "정보기술", "반도체"), ("LRCX", "Lam Research", "정보기술", "반도체장비"),
        ("KLAC", "KLA Corp", "정보기술", "반도체장비"), ("AMAT", "Applied Materials", "정보기술", "반도체장비"),
        ("SNPS", "Synopsys", "정보기술", "반도체설계"), ("CDNS", "Cadence", "정보기술", "반도체설계"),
        ("ARM", "Arm Holdings", "정보기술", "반도체설계"), ("ON", "ON Semi", "정보기술", "반도체"),
        ("NFLX", "Netflix", "통신서비스", "스트리밍"), ("ADBE", "Adobe", "정보기술", "소프트웨어"),
        ("ORCL", "Oracle", "정보기술", "소프트웨어"), ("CSCO", "Cisco", "정보기술", "네트워크"),
        ("INTU", "Intuit", "정보기술", "소프트웨어"), ("NOW", "ServiceNow", "정보기술", "소프트웨어"),
        ("PLTR", "Palantir", "정보기술", "데이터분석"), ("SNOW", "Snowflake", "정보기술", "클라우드"),
        ("PANW", "Palo Alto Networks", "정보기술", "사이버보안"), ("CRWD", "CrowdStrike", "정보기술", "사이버보안"),
        ("UBER", "Uber", "산업재", "모빌리티"), ("ABNB", "Airbnb", "경기소비재", "여행"),
        ("COIN", "Coinbase", "금융", "암호화폐"), ("SQ", "Block", "금융", "핀테크"),
        ("PYPL", "PayPal", "금융", "핀테크"), ("SHOP", "Shopify", "정보기술", "이커머스"),
        ("COST", "Costco", "필수소비재", "유통"), ("SBUX", "Starbucks", "경기소비재", "외식"),
        ("GILD", "Gilead", "헬스케어", "바이오"), ("AMGN", "Amgen", "헬스케어", "바이오"),
        ("ISRG", "Intuitive Surgical", "헬스케어", "의료기기"), ("REGN", "Regeneron", "헬스케어", "바이오"),
        ("MRNA", "Moderna", "헬스케어", "바이오"), ("DXCM", "DexCom", "헬스케어", "의료기기"),
        ("CME", "CME Group", "금융", "거래소"), ("ICE", "ICE", "금융", "거래소"),
        ("BLK", "BlackRock", "금융", "자산운용"), ("SCHW", "Schwab", "금융", "증권"),
        ("MS", "Morgan Stanley", "금융", "투자은행"), ("C", "Citigroup", "금융", "은행"),
        ("BAC", "Bank of America", "금융", "은행"), ("WFC", "Wells Fargo", "금융", "은행"),
        ("T", "AT&T", "통신서비스", "통신"), ("VZ", "Verizon", "통신서비스", "통신"),
        ("DIS", "Disney", "통신서비스", "미디어"), ("CMCSA", "Comcast", "통신서비스", "미디어"),
        ("NKE", "Nike", "경기소비재", "스포츠"), ("LOW", "Lowe's", "경기소비재", "주택개선"),
        ("UPS", "UPS", "산업재", "물류"), ("RTX", "RTX", "산업재", "방산"),
        ("LMT", "Lockheed Martin", "산업재", "방산"), ("GE", "GE Aerospace", "산업재", "항공우주"),
        ("DE", "Deere", "산업재", "농기계"), ("HON", "Honeywell", "산업재", "복합"),
        ("COP", "ConocoPhillips", "에너지", "석유"), ("SLB", "SLB", "에너지", "유전서비스"),
        ("LIN", "Linde", "소재", "산업가스"), ("APD", "Air Products", "소재", "산업가스"),
        ("SO", "Southern Co", "유틸리티", "전력"), ("DUK", "Duke Energy", "유틸리티", "전력"),
    ],
    "kr": [
        ("005930.KS", "삼성전자", "정보기술", "반도체"), ("000660.KS", "SK하이닉스", "정보기술", "반도체"),
        ("035420.KS", "NAVER", "통신서비스", "인터넷"), ("005380.KS", "현대차", "경기소비재", "자동차"),
        ("035720.KS", "카카오", "통신서비스", "인터넷"), ("051910.KS", "LG화학", "소재", "화학"),
        ("006400.KS", "삼성SDI", "정보기술", "배터리"), ("068270.KS", "셀트리온", "헬스케어", "바이오"),
        ("105560.KS", "KB금융", "금융", "은행"), ("055550.KS", "신한지주", "금융", "은행"),
        ("066570.KS", "LG전자", "경기소비재", "가전"), ("012330.KS", "현대모비스", "경기소비재", "자동차부품"),
        ("207940.KS", "삼성바이오", "헬스케어", "바이오"), ("000270.KS", "기아", "경기소비재", "자동차"),
        ("096770.KS", "SK이노베이션", "에너지", "정유"), ("034730.KS", "SK", "산업재", "지주회사"),
        ("030200.KS", "KT", "통신서비스", "통신"), ("017670.KS", "SK텔레콤", "통신서비스", "통신"),
        ("003670.KS", "포스코홀딩스", "소재", "철강"), ("010950.KS", "S-Oil", "에너지", "정유"),
        ("086790.KS", "하나금융", "금융", "은행"), ("316140.KS", "우리금융", "금융", "은행"),
        ("047810.KS", "한국항공우주", "산업재", "방산"), ("009150.KS", "삼성전기", "정보기술", "전자부품"),
        ("033780.KS", "KT&G", "필수소비재", "담배"), ("028260.KS", "삼성물산", "산업재", "건설"),
        ("373220.KS", "LG에너지솔루션", "정보기술", "배터리"), ("018260.KS", "삼성SDS", "정보기술", "IT서비스"),
        ("032830.KS", "삼성생명", "금융", "보험"), ("003550.KS", "LG", "산업재", "지주회사"),
    ],
    "jp": [
        ("7203.T", "Toyota", "경기소비재", "자동차"), ("6758.T", "Sony", "경기소비재", "전자"),
        ("9984.T", "SoftBank Group", "통신서비스", "투자"), ("6861.T", "Keyence", "정보기술", "센서"),
        ("8306.T", "MUFG", "금융", "은행"), ("9433.T", "KDDI", "통신서비스", "통신"),
        ("6501.T", "Hitachi", "산업재", "전자"), ("4063.T", "Shin-Etsu", "소재", "화학"),
        ("7267.T", "Honda", "경기소비재", "자동차"), ("8035.T", "Tokyo Electron", "정보기술", "반도체장비"),
        ("7741.T", "HOYA", "헬스케어", "의료기기"), ("4502.T", "Takeda", "헬스케어", "제약"),
        ("7974.T", "Nintendo", "통신서비스", "게임"), ("6594.T", "Nidec", "산업재", "모터"),
        ("8766.T", "Tokio Marine", "금융", "보험"), ("9432.T", "NTT", "통신서비스", "통신"),
        ("6981.T", "Murata", "정보기술", "전자부품"), ("4568.T", "Daiichi Sankyo", "헬스케어", "제약"),
        ("6367.T", "Daikin", "산업재", "공조"), ("6902.T", "Denso", "경기소비재", "자동차부품"),
        ("8001.T", "ITOCHU", "산업재", "종합상사"), ("8058.T", "Mitsubishi Corp", "산업재", "종합상사"),
        ("6857.T", "Advantest", "정보기술", "반도체장비"), ("9983.T", "Fast Retailing", "경기소비재", "의류"),
        ("6098.T", "Recruit", "산업재", "인력"), ("4661.T", "Oriental Land", "경기소비재", "레저"),
        ("6723.T", "Renesas", "정보기술", "반도체"), ("4519.T", "Chugai", "헬스케어", "제약"),
        ("6273.T", "SMC", "산업재", "자동화"), ("7832.T", "Bandai Namco", "통신서비스", "엔터"),
    ],
    "eu": [
        ("ASML.AS", "ASML", "정보기술", "반도체장비"), ("MC.PA", "LVMH", "경기소비재", "명품"),
        ("SAP.DE", "SAP", "정보기술", "소프트웨어"), ("SIE.DE", "Siemens", "산업재", "복합"),
        ("OR.PA", "L'Oreal", "필수소비재", "화장품"), ("TTE.PA", "TotalEnergies", "에너지", "석유"),
        ("SAN.PA", "Sanofi", "헬스케어", "제약"), ("AIR.PA", "Airbus", "산업재", "항공우주"),
        ("BAS.DE", "BASF", "소재", "화학"), ("DTE.DE", "Deutsche Telekom", "통신서비스", "통신"),
        ("ADS.DE", "Adidas", "경기소비재", "스포츠"), ("BNP.PA", "BNP Paribas", "금융", "은행"),
        ("AI.PA", "Air Liquide", "소재", "산업가스"), ("SU.PA", "Schneider", "산업재", "전기"),
        ("RMS.PA", "Hermes", "경기소비재", "명품"), ("DG.PA", "Vinci", "산업재", "건설"),
        ("PHIA.AS", "Philips", "헬스케어", "의료기기"), ("INGA.AS", "ING", "금융", "은행"),
        ("NESN.SW", "Nestle", "필수소비재", "식품"), ("ROG.SW", "Roche", "헬스케어", "제약"),
        ("NOVN.SW", "Novartis", "헬스케어", "제약"), ("ABI.BR", "AB InBev", "필수소비재", "맥주"),
        ("SHEL.L", "Shell", "에너지", "석유"), ("AZN.L", "AstraZeneca", "헬스케어", "제약"),
        ("ULVR.L", "Unilever", "필수소비재", "생활용품"), ("BARC.L", "Barclays", "금융", "은행"),
        ("GSK.L", "GSK", "헬스케어", "제약"), ("NOKIA.HE", "Nokia", "정보기술", "통신장비"),
        ("VOW3.DE", "Volkswagen", "경기소비재", "자동차"), ("BMW.DE", "BMW", "경기소비재", "자동차"),
    ],
}


# ===== 확장 종목 리스트 (전체 시장 검색용) =====
EXPANDED_STOCKS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "expanded_stocks.json")
_expanded_cache = {}
_expanded_time = 0
_top_gainers_cache = {}


def _fetch_us_all_tickers():
    """미국 전체 상장 종목 (NASDAQ + NYSE)."""
    us_map = load_us_stock_map()
    return list(us_map.keys())


def _fetch_kr_tickers():
    """한국 상장 종목 (DART 기업 맵 기반, KOSPI .KS)."""
    corp_map = load_dart_corp_map()
    tickers = []
    for code in corp_map:
        if len(code) == 6 and code.isdigit():
            tickers.append(code + ".KS")
    return tickers[:500]


def _fetch_nikkei225_tickers():
    """Nikkei 225 종목 (Wikipedia)."""
    tables = pd.read_html("https://en.wikipedia.org/wiki/Nikkei_225")
    for table in tables:
        for col in table.columns:
            try:
                vals = table[col].astype(str).tolist()
                codes = [v.strip() + ".T" for v in vals
                         if v.strip().isdigit() and len(v.strip()) == 4]
                if len(codes) >= 50:
                    return codes
            except Exception:
                continue
    return []


def _fetch_eu_tickers():
    """유럽 주요 종목 (FTSE 100 + DAX 40 + 기존 목록)."""
    tickers = set()
    for s in MARKET_STOCKS.get("eu", []):
        tickers.add(s[0])
    # FTSE 100
    try:
        tables = pd.read_html("https://en.wikipedia.org/wiki/FTSE_100_Index")
        for table in tables:
            for col in table.columns:
                if "ticker" in str(col).lower() or "epic" in str(col).lower():
                    for v in table[col]:
                        t = str(v).strip()
                        if t and t != "nan" and len(t) <= 5:
                            tickers.add(t + ".L" if "." not in t else t)
                    break
    except Exception:
        pass
    # DAX 40
    try:
        tables = pd.read_html("https://en.wikipedia.org/wiki/DAX")
        for table in tables:
            for col in table.columns:
                if "ticker" in str(col).lower() or "symbol" in str(col).lower():
                    for v in table[col]:
                        t = str(v).strip()
                        if t and t != "nan" and len(t) <= 6:
                            tickers.add(t + ".DE" if "." not in t else t)
                    break
    except Exception:
        pass
    return list(tickers)


def load_expanded_tickers(market):
    """시장별 확장 종목 리스트 로드 (디스크 캐시 7일)."""
    global _expanded_cache, _expanded_time

    now = time.time()
    if market in _expanded_cache and now - _expanded_time < 86400:
        return _expanded_cache[market]

    # 디스크 캐시
    try:
        mod_time = os.path.getmtime(EXPANDED_STOCKS_FILE)
        if now - mod_time < 86400 * 7:
            with open(EXPANDED_STOCKS_FILE, "r", encoding="utf-8") as f:
                _expanded_cache = json.load(f)
                _expanded_time = now
                if market in _expanded_cache and _expanded_cache[market]:
                    return _expanded_cache[market]
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        pass

    # 새로 가져오기
    fetchers = {
        "us": _fetch_us_all_tickers,
        "kr": _fetch_kr_tickers,
        "jp": _fetch_nikkei225_tickers,
        "eu": _fetch_eu_tickers,
    }
    fetcher = fetchers.get(market)
    if fetcher:
        try:
            result = fetcher()
            if result:
                _expanded_cache[market] = result
        except Exception as e:
            print(f"[Expanded] {market} fetch failed: {e}")

    # 폴백: MARKET_STOCKS
    if market not in _expanded_cache or not _expanded_cache[market]:
        _expanded_cache[market] = [s[0] for s in MARKET_STOCKS.get(market, [])]

    _expanded_time = now
    try:
        with open(EXPANDED_STOCKS_FILE, "w", encoding="utf-8") as f:
            json.dump(_expanded_cache, f, ensure_ascii=False)
    except Exception:
        pass

    return _expanded_cache.get(market, [])


def _download_closes_chunked(tickers, yf_period, chunk_size=500):
    """대량 종목 Close 가격을 청크 단위로 다운로드."""
    all_closes = None
    for i in range(0, len(tickers), chunk_size):
        chunk = tickers[i:i + chunk_size]
        try:
            data = yf.download(chunk, period=yf_period, progress=False, threads=False)
            if data.empty:
                continue
            closes = data["Close"]
            if isinstance(closes, pd.Series):
                closes = closes.to_frame(name=chunk[0])
            if all_closes is None:
                all_closes = closes
            else:
                all_closes = pd.concat([all_closes, closes], axis=1)
        except Exception:
            continue
    return all_closes


def _calc_change(series, period):
    """기간에 맞는 변동률 계산."""
    if period == "1d":
        # 직전 1거래일 변동: 전일 종가 → 당일 종가
        return (series.iloc[-1] - series.iloc[-2]) / series.iloc[-2] * 100
    else:
        # 전체 기간 변동: 시작 → 현재
        return (series.iloc[-1] - series.iloc[0]) / series.iloc[0] * 100


_CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cache")


@app.route("/api/top-gainers", methods=["GET"])
def get_top_gainers():
    """시장별 상승률 TOP 10 (배치 프리페치 파일 기반)"""
    market = request.args.get("market", "us")
    period = request.args.get("period", "1w")

    if market not in ("us", "kr", "jp", "eu"):
        return jsonify({"error": "지원하지 않는 시장입니다"}), 404

    cache_path = os.path.join(_CACHE_DIR, f"top_gainers_{market}.json")
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            cached = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return jsonify({"error": "아직 배치가 실행되지 않았습니다"}), 503

    items = cached.get("data", {}).get(period, [])
    return jsonify({
        "data": items,
        "last_updated": cached.get("last_updated", ""),
        "universe_size": cached.get("universe_size", 0),
        "failed_count": cached.get("failed_count", 0),
    })


def _is_private_ip(addr):
    import ipaddress
    try:
        ip = ipaddress.ip_address(addr)
        if ip.is_loopback or ip.is_private:
            return True
        # Tailscale CGNAT: 100.64.0.0/10
        return ipaddress.ip_address("100.64.0.0") <= ip <= ipaddress.ip_address("100.127.255.255")
    except ValueError:
        return False


@app.route("/api/top-gainers/refresh", methods=["POST"])
def refresh_top_gainers():
    if not _is_private_ip(request.remote_addr):
        return jsonify({"error": "forbidden"}), 403
    import subprocess, sys
    market = request.args.get("market", "us")
    subprocess.Popen(
        [sys.executable, os.path.join(os.path.dirname(os.path.abspath(__file__)), "batch_top_gainers.py"), market],
        cwd=os.path.dirname(os.path.abspath(__file__)),
    )
    return jsonify({"status": "triggered"})


@app.route("/api/sector-leaders", methods=["GET"])
def get_sector_leaders():
    cache_path = os.path.join(_CACHE_DIR, "sector_leaders.json")
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            cached = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return jsonify({"error": "아직 배치가 실행되지 않았습니다"}), 503
    return jsonify(cached)


@app.route("/api/stock-leaders", methods=["GET"])
def get_stock_leaders():
    cache_path = os.path.join(_CACHE_DIR, "stock_leaders.json")
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            cached = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return jsonify({"error": "아직 배치가 실행되지 않았습니다"}), 503
    return jsonify(cached)


@app.route("/api/stock-leaders/custom", methods=["GET"])
def get_stock_leaders_custom():
    from datetime import datetime, timedelta
    start = request.args.get("start", "")
    end = request.args.get("end", "")
    if not start or not end:
        return jsonify({"error": "start and end parameters required"}), 400
    try:
        dt_s = datetime.strptime(start, "%Y-%m-%d")
        dt_e = datetime.strptime(end, "%Y-%m-%d")
    except ValueError:
        return jsonify({"error": "invalid date format, use YYYY-MM-DD"}), 400
    if dt_s >= dt_e:
        return jsonify({"error": "start must be before end"}), 400
    oldest = datetime.now() - timedelta(days=504)
    if dt_s < oldest:
        return jsonify({"error": f"start date too old, earliest allowed: {oldest.strftime('%Y-%m-%d')}"}), 400
    if dt_e > datetime.now():
        return jsonify({"error": "end date cannot be in the future"}), 400
    import stock_leaders
    result = stock_leaders.compute_stock_leaders_custom(start, end)
    if "error" in result:
        return jsonify(result), 503
    return jsonify(result)


@app.route("/api/valuation", methods=["GET"])
def get_valuation():
    cache_path = os.path.join(_CACHE_DIR, "valuation_screener.json")
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            cached = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return jsonify({"error": "아직 배치가 실행되지 않았습니다"}), 503
    return jsonify(cached)


@app.route("/api/valuation/band/<ticker>", methods=["GET"])
def get_valuation_band(ticker):
    cache_path = os.path.join(_CACHE_DIR, "valuation_bands.json")
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return jsonify({"error": "아직 배치가 실행되지 않았습니다"}), 503
    bands = data.get("bands", {})
    entry = bands.get(ticker)
    if not entry:
        return jsonify({"error": "밴드 없음"}), 404
    return jsonify(entry)


@app.route("/api/valuation/overview/<code>", methods=["GET"])
def get_valuation_overview(code):
    name = request.args.get("name", "")
    try:
        corp_code = earnings_tracker.get_corp_code(code)
    except Exception:
        corp_code = None
    if not corp_code:
        return jsonify({"overview": None})
    try:
        overview = earnings_tracker.get_business_overview_cached(code, name, corp_code)
    except Exception:
        overview = None
    return jsonify({"overview": overview})


@app.route("/api/sector-leaders/custom", methods=["GET"])
def get_sector_leaders_custom():
    from datetime import datetime
    start = request.args.get("start", "")
    end = request.args.get("end", "")
    if not start or not end:
        return jsonify({"error": "start와 end 파라미터가 필요합니다"}), 400
    try:
        dt_s = datetime.strptime(start, "%Y-%m-%d")
        dt_e = datetime.strptime(end, "%Y-%m-%d")
    except ValueError:
        return jsonify({"error": "날짜 형식이 잘못되었습니다 (YYYY-MM-DD)"}), 400
    if dt_s >= dt_e:
        return jsonify({"error": "시작일이 종료일보다 앞서야 합니다"}), 400
    import sector_leaders
    bounds = sector_leaders.get_custom_bounds()
    oldest = bounds.get("oldest")
    newest = bounds.get("newest")
    if oldest and start < oldest:
        return jsonify({"error": f"시작일이 캐시 범위({oldest}~)를 벗어났습니다", "bounds": bounds}), 400
    if newest and end > newest:
        return jsonify({"error": f"종료일이 캐시 범위(~{newest})를 벗어났습니다", "bounds": bounds}), 400
    result = sector_leaders.compute_sector_leaders_custom(start, end)
    if "error" in result:
        return jsonify(result), 503
    result["bounds"] = bounds
    return jsonify(result)


@app.route("/api/sector-leaders/bounds", methods=["GET"])
def get_sector_leaders_bounds():
    import sector_leaders
    return jsonify(sector_leaders.get_custom_bounds())


@app.route("/api/sector-leaders/<region>", methods=["GET"])
def get_sector_leaders_by_region(region):
    if region not in ("kr", "us"):
        return jsonify({"error": "지원하지 않는 지역입니다 (kr, us)"}), 400
    cache_path = os.path.join(_CACHE_DIR, "sector_leaders.json")
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            cached = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return jsonify({"error": "아직 배치가 실행되지 않았습니다"}), 503
    result = cached.get(region, {})
    result["last_updated"] = cached.get("last_updated", "")
    return jsonify(result)


@app.route("/api/sector-leaders/refresh", methods=["POST"])
def refresh_sector_leaders():
    if not _is_private_ip(request.remote_addr):
        return jsonify({"error": "forbidden"}), 403
    import subprocess, sys
    region = request.args.get("region", "all")
    if region not in ("kr", "us", "all"):
        return jsonify({"error": "지원하지 않는 지역입니다 (kr, us, all)"}), 400
    subprocess.Popen(
        [sys.executable, os.path.join(os.path.dirname(os.path.abspath(__file__)), "sector_leaders.py"), region],
        cwd=os.path.dirname(os.path.abspath(__file__)),
    )
    return jsonify({"status": "triggered"}), 202


# 한국어 → 영어 키워드 매핑
KO_TO_EN = {
    "반도체": "semiconductor",
    "환율": "exchange rate USD KRW",
    "금리": "interest rate",
    "인플레이션": "inflation",
    "경기침체": "recession",
    "부동산": "real estate",
    "원유": "crude oil",
    "금": "gold",
    "은행": "banking",
    "주식": "stock market",
    "채권": "bond",
    "무역": "trade",
    "관세": "tariff",
    "전쟁": "war",
    "AI": "artificial intelligence",
    "인공지능": "artificial intelligence",
    "암호화폐": "cryptocurrency",
    "비트코인": "bitcoin",
    "테슬라": "Tesla",
    "애플": "Apple",
    "엔비디아": "Nvidia",
    "삼성": "Samsung",
    "중국": "China",
    "일본": "Japan",
    "유럽": "Europe",
}

TRUSTED_SOURCES = [
    "reuters", "bloomberg", "wsj", "wall street journal",
    "cnbc", "financial times", "ft.com", "barron",
    "marketwatch", "economist", "associated press", "ap news",
]

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")
claude_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)


def fetch_rss_articles(search_terms):
    """구글 뉴스 RSS에서 신뢰 소스 기사를 수집한다. 12시간 이내 기사만."""
    from email.utils import parsedate_to_datetime
    articles = []
    seen_titles = set()
    cutoff = datetime.now().astimezone() - timedelta(hours=12)

    for term in search_terms:
        try:
            url = ("https://news.google.com/rss/search?q="
                   + requests.utils.quote(term)
                   + "&hl=en-US&gl=US&ceid=US:en")
            feed = feedparser.parse(url)
            for entry in feed.entries:
                source = (entry.get("source", {}).get("title", "")
                          if hasattr(entry, "source") else "")
                source_lower = source.lower()
                if not any(t in source_lower for t in TRUSTED_SOURCES):
                    continue
                title = entry.get("title", "")
                if title in seen_titles:
                    continue
                published = entry.get("published", "")
                if published:
                    try:
                        pub_dt = parsedate_to_datetime(published)
                        if pub_dt < cutoff:
                            continue
                    except Exception:
                        pass
                seen_titles.add(title)
                articles.append({
                    "title": title,
                    "source": source,
                    "url": entry.get("link", ""),
                    "publishedAt": published,
                })
        except Exception:
            continue

    articles.sort(key=lambda a: a.get("publishedAt", ""), reverse=True)
    return articles[:20]


KO_TRUSTED_SOURCES = [
    "한국경제", "매일경제", "연합뉴스", "조선비즈",
    "머니투데이", "이데일리", "서울경제", "한겨레",
    "동아일보", "중앙일보", "뉴스1", "뉴시스",
]

EN_STOCK_TRUSTED_SOURCES = [
    "reuters", "bloomberg", "wsj", "wall street journal",
    "cnbc", "barron", "marketwatch", "financial times", "ft.com",
    "associated press", "ap news", "economist",
    "yahoo finance", "seeking alpha", "marketbeat", "benzinga",
    "investing.com", "motley fool", "investor's business daily",
    "morningstar", "investopedia", "fortune", "business insider",
    "24/7 wall st", "barchart", "thestreet", "zacks",
]


def fetch_stock_news(query, lang="en", max_results=10):
    """종목 관련 뉴스를 언어별로 검색."""
    from email.utils import parsedate_to_datetime

    if lang == "ko":
        url = ("https://news.google.com/rss/search?q="
               + requests.utils.quote(query)
               + "&hl=ko&gl=KR&ceid=KR:ko")
        trusted = KO_TRUSTED_SOURCES
    else:
        url = ("https://news.google.com/rss/search?q="
               + requests.utils.quote(query)
               + "&hl=en-US&gl=US&ceid=US:en")
        trusted = EN_STOCK_TRUSTED_SOURCES

    cutoff = datetime.now().astimezone() - timedelta(hours=24)
    articles = []
    seen = set()

    try:
        feed = feedparser.parse(url)
        for entry in feed.entries:
            source = (entry.get("source", {}).get("title", "")
                      if hasattr(entry, "source") else "")
            if not any(t.lower() in source.lower() for t in trusted):
                continue
            title = entry.get("title", "")
            if title in seen:
                continue
            published = entry.get("published", "")
            if published:
                try:
                    if parsedate_to_datetime(published) < cutoff:
                        continue
                except Exception:
                    pass
            seen.add(title)
            articles.append({
                "title": title,
                "source": source,
                "url": entry.get("link", ""),
                "publishedAt": published,
            })
    except Exception:
        return []

    articles.sort(key=lambda a: a.get("publishedAt", ""), reverse=True)
    return articles[:max_results]


def rank_group_articles(group_name, articles, top_n=3):
    """Claude Haiku로 그룹 내 기사를 금융시장 영향력 기준으로 선별하고 3줄 요약한다."""
    if not articles:
        return []

    article_list = "\n".join(
        f"{i+1}. [{a['source']}] {a['title']}"
        for i, a in enumerate(articles)
    )

    response = claude_client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=2048,
        messages=[{
            "role": "user",
            "content": (
                f"주제: \"{group_name}\"\n\n"
                f"다음은 이 주제와 관련된 글로벌 금융 뉴스 목록이다.\n"
                f"금융시장(주식, 채권, 환율, 원자재)에 미치는 영향력이 큰 순서대로 "
                f"상위 {top_n}개를 선별해라. 중복되거나 사소한 뉴스는 제외해라.\n\n"
                "각 뉴스에 대해:\n"
                "1. 제목을 바탕으로 뉴스의 핵심 내용을 한국어 3줄로 요약해라.\n"
                "2. 각 줄은 '- '로 시작하는 완결된 문장이어야 한다.\n"
                "3. 금융시장에 미치는 영향을 반드시 포함해라.\n\n"
                f"{article_list}\n\n"
                "반드시 아래 JSON 배열 형식으로만 응답해라. 다른 텍스트는 절대 포함하지 마라.\n"
                '[{"rank":1,"index":원래번호,"summary":["요약1","요약2","요약3"]},...]\n'
                "index는 위 목록의 번호(1부터 시작)이다."
            ),
        }],
    )

    raw = response.content[0].text.strip()
    start = raw.find("[")
    end = raw.rfind("]") + 1
    if start == -1 or end == 0:
        return articles[:top_n]

    ranked = json.loads(raw[start:end])
    result = []
    for item in ranked[:top_n]:
        idx = item["index"] - 1
        if 0 <= idx < len(articles):
            a = articles[idx].copy()
            a["summary"] = item.get("summary", [])
            result.append(a)
    return result


@app.route("/api/news", methods=["POST"])
def get_news():
    """그룹별 구글 RSS + Claude Haiku 기반 금융 뉴스 선별"""
    body = request.get_json()
    if not body:
        return jsonify([])

    groups = body.get("groups", [])
    if not groups:
        return jsonify([])

    results = []
    for group in groups:
        name = group.get("name", "")
        terms = group.get("terms", [])
        if not terms:
            continue

        # 한국어 검색어를 영어로 변환
        translated = [KO_TO_EN.get(t, t) for t in terms]

        articles = fetch_rss_articles(translated)
        if not articles:
            results.append({"group": name, "articles": []})
            continue

        try:
            ranked = rank_group_articles(name, articles, top_n=3)
            results.append({"group": name, "articles": ranked})
        except Exception:
            results.append({"group": name, "articles": articles[:3]})

    return jsonify(results)


KEYWORDS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "keywords.json")

DEFAULT_GROUPS = [
    {"name": "미국 금리/Fed", "terms": ["Fed interest rate", "FOMC", "Federal Reserve policy"]},
    {"name": "반도체 산업", "terms": ["semiconductor", "Nvidia earnings", "chip export"]},
    {"name": "국제 유가", "terms": ["crude oil price", "OPEC production", "WTI Brent"]},
]


def load_keywords():
    try:
        with open(KEYWORDS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        save_keywords(DEFAULT_GROUPS)
        return DEFAULT_GROUPS


def save_keywords(groups):
    with open(KEYWORDS_FILE, "w", encoding="utf-8") as f:
        json.dump(groups, f, ensure_ascii=False, indent=2)


@app.route("/api/keywords", methods=["GET"])
def get_keywords():
    return jsonify(load_keywords())


@app.route("/api/keywords", methods=["POST"])
def set_keywords():
    groups = request.get_json()
    if not isinstance(groups, list):
        return jsonify({"error": "groups must be a list"}), 400
    save_keywords(groups)
    return jsonify({"ok": True})


# ===== 캘린더 기능 =====

CALENDAR_WATCHLIST_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "calendar_watchlist.json")

DEFAULT_CAL_WATCHLIST = {
    "us_earnings": ["AAPL", "MSFT", "GOOGL", "AMZN", "NVDA", "META", "TSLA"],
    "kr_earnings": ["005930", "000660", "035420"],
    "dividends": ["AAPL", "MSFT", "JNJ", "KO", "PG"],
}

# ===== US 전체 종목 리스트 (나스닥 + NYSE 등) =====
_US_STOCK_MAP_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "us_stock_map.json")
_us_stock_map = {}
_us_stock_map_time = 0


def load_us_stock_map():
    """나스닥+NYSE 전체 종목 리스트 로드 (디스크 캐시 7일)"""
    global _us_stock_map, _us_stock_map_time
    now = time.time()
    if _us_stock_map and now - _us_stock_map_time < 86400 * 7:
        return _us_stock_map

    # 디스크 캐시 확인
    try:
        mod_time = os.path.getmtime(_US_STOCK_MAP_FILE)
        if now - mod_time < 86400 * 7:
            with open(_US_STOCK_MAP_FILE, "r", encoding="utf-8") as f:
                _us_stock_map = json.load(f)
                _us_stock_map_time = now
                return _us_stock_map
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        pass

    # 다운로드
    mapping = {}
    for url, exchange in [
        ("https://www.nasdaqtrader.com/dynamic/SymDir/nasdaqlisted.txt", "NASDAQ"),
        ("https://www.nasdaqtrader.com/dynamic/SymDir/otherlisted.txt", "NYSE"),
    ]:
        try:
            res = requests.get(url, timeout=15)
            lines = res.text.strip().split("\n")
            for line in lines[1:]:  # 헤더 스킵
                parts = line.split("|")
                if len(parts) < 2:
                    continue
                symbol = parts[0].strip()
                if exchange == "NASDAQ":
                    name = parts[1].strip()  # Security Name
                else:
                    name = parts[1].strip()  # Security Name
                    symbol = parts[7].strip() if len(parts) > 7 and parts[7].strip() else symbol  # NASDAQ symbol
                # 유효한 티커만 (알파벳, 테스트 종목 제외)
                if (not symbol or not name or len(symbol) > 5
                        or "$" in symbol or "File Creation" in name
                        or symbol.endswith("W") or symbol.endswith("R")):
                    continue
                mapping[symbol] = {"name": name, "exchange": exchange}
        except Exception:
            continue

    if mapping:
        _us_stock_map = mapping
        _us_stock_map_time = now
        with open(_US_STOCK_MAP_FILE, "w", encoding="utf-8") as f:
            json.dump(mapping, f, ensure_ascii=False)
    return _us_stock_map


# DART API
DART_API_KEY = os.environ.get("DART_API_KEY")
DART_CORP_MAP_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dart_corp_map.json")
_dart_corp_map = {}
_dart_corp_map_time = 0


def load_dart_corp_map():
    """DART 고유번호(corp_code) 매핑 로드 (디스크 캐시 + 주간 갱신)"""
    global _dart_corp_map, _dart_corp_map_time
    now = time.time()
    if _dart_corp_map and now - _dart_corp_map_time < 86400 * 7:
        return _dart_corp_map

    # 디스크 캐시 확인
    try:
        mod_time = os.path.getmtime(DART_CORP_MAP_FILE)
        if now - mod_time < 86400 * 7:
            with open(DART_CORP_MAP_FILE, "r", encoding="utf-8") as f:
                _dart_corp_map = json.load(f)
                _dart_corp_map_time = now
                return _dart_corp_map
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        pass

    # DART에서 새로 다운로드
    try:
        res = requests.get(
            "https://opendart.fss.or.kr/api/corpCode.xml",
            params={"crtfc_key": DART_API_KEY},
            timeout=30,
        )
        with zipfile.ZipFile(io.BytesIO(res.content)) as z:
            with z.open("CORPCODE.xml") as f:
                tree = ET.parse(f)
                root = tree.getroot()
                mapping = {}
                for item in root.iter("list"):
                    sc = (item.findtext("stock_code") or "").strip()
                    cc = (item.findtext("corp_code") or "").strip()
                    cn = (item.findtext("corp_name") or "").strip()
                    if sc and cc:
                        mapping[sc] = {"corp_code": cc, "name": cn}
                _dart_corp_map = mapping
                _dart_corp_map_time = now
        with open(DART_CORP_MAP_FILE, "w", encoding="utf-8") as f:
            json.dump(_dart_corp_map, f, ensure_ascii=False)
    except Exception:
        pass
    return _dart_corp_map


def fetch_dart_events(stock_code, year, month):
    """DART API에서 특정 종목의 실적공시/IR/주주총회 검색"""
    corp_map = load_dart_corp_map()
    clean = stock_code.replace(".KS", "").replace(".KQ", "")
    info = corp_map.get(clean)
    if not info:
        return []

    corp_code = info["corp_code"]
    corp_name = info["name"]
    bgn = f"{year}{month:02d}01"
    last = (date(year, month, 1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    end = last.strftime("%Y%m%d")

    events = []
    try:
        res = requests.get(
            "https://opendart.fss.or.kr/api/list.json",
            params={
                "crtfc_key": DART_API_KEY,
                "corp_code": corp_code,
                "bgn_de": bgn,
                "end_de": end,
                "page_count": 100,
            },
            timeout=10,
        )
        data = res.json()
        if data.get("status") != "000":
            return []

        for item in data.get("list", []):
            title = item.get("report_nm", "")
            dt_raw = item.get("rcept_dt", "")
            if len(dt_raw) != 8:
                continue
            ds = f"{dt_raw[:4]}-{dt_raw[4:6]}-{dt_raw[6:8]}"

            is_earnings = any(k in title for k in ["사업보고서", "분기보고서", "반기보고서"])
            is_ir = any(k in title for k in ["투자설명", "IR", "기업설명회"])
            is_agm = "주주총회" in title

            if is_earnings:
                events.append({"date": ds, "title": f"{corp_name} 실적공시", "detail": title})
            elif is_ir:
                events.append({"date": ds, "title": f"{corp_name} IR", "detail": title})
            elif is_agm:
                events.append({"date": ds, "title": f"{corp_name} 주주총회", "detail": title})
    except Exception:
        pass
    return events

# 미국 경제지표 발표일 (월별 일자, 인덱스 0=1월)
US_ECONOMIC_SCHEDULE = {
    2025: {
        "CPI":       [15, 12, 12, 10, 13, 11, 10, 12, 10, 14, 12, 10],
        "PPI":       [14, 13, 13, 11, 15, 12, 15, 14, 11, 15, 13, 11],
        "고용보고서(NFP)": [10,  7,  7,  4,  2,  6,  3,  1,  5,  3,  7,  5],
        "PCE":       [31, 28, 28, 30, 30, 27, 31, 29, 26, 31, 26, 23],
        "GDP":       [30, 27, 27, 30, 29, 26, 30, 28, 25, 29, 26, 23],
    },
    2026: {
        "CPI":       [13, 11, 11, 14, 12, 10, 14, 12, 15, 13, 10,  9],
        "PPI":       [15, 12, 12,  9, 14, 11, 16, 13, 10, 15, 12, 10],
        "고용보고서(NFP)": [ 9,  6,  6,  3,  8,  5,  2,  7,  4,  2,  6,  4],
        "PCE":       [30, 27, 27, 30, 29, 26, 31, 28, 25, 30, 25, 23],
        "GDP":       [29, 26, 26, 29, 28, 25, 30, 27, 24, 29, 25, 22],
    },
}

# FOMC 금리결정일 (회의 둘째 날)
FOMC_DATES = {
    2025: [(1,29),(3,19),(5,7),(6,18),(7,30),(9,17),(10,29),(12,10)],
    2026: [(1,28),(3,18),(4,29),(6,17),(7,29),(9,16),(10,28),(12,9)],
}

ECB_DATES = {
    2025: [(1,30),(3,6),(4,17),(6,5),(7,24),(9,11),(10,30),(12,18)],
    2026: [(1,22),(3,5),(4,16),(6,4),(7,16),(9,10),(10,29),(12,17)],
}

BOJ_DATES = {
    2025: [(1,24),(3,14),(5,1),(6,17),(7,31),(9,19),(10,30),(12,19)],
    2026: [(1,16),(3,13),(4,28),(6,16),(7,15),(9,18),(10,29),(12,18)],
}

# 한국은행 금통위 금리결정일
BOK_DATES = {
    2025: [(1,16),(2,27),(4,17),(5,29),(7,10),(8,21),(10,16),(11,27)],
    2026: [(1,15),(2,26),(4,16),(5,28),(7,9),(8,27),(10,15),(11,26)],
}

# 한국 CPI 발표일
KR_CPI_DATES = {
    2025: [(1,3),(2,4),(3,4),(4,1),(5,6),(6,3),(7,1),(8,5),(9,2),(10,2),(11,4),(12,2)],
    2026: [(1,2),(2,3),(3,3),(4,1),(5,5),(6,2),(7,1),(8,4),(9,1),(10,1),(11,3),(12,1)],
}

_cal_cache = {}
_cal_cache_ttl = 3600


def load_cal_watchlist():
    try:
        with open(CALENDAR_WATCHLIST_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        save_cal_watchlist(DEFAULT_CAL_WATCHLIST)
        return DEFAULT_CAL_WATCHLIST


def save_cal_watchlist(data):
    global _cal_cache
    with open(CALENDAR_WATCHLIST_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    _cal_cache = {}


def get_third_friday(year, month):
    """셋째 금요일 계산 (네 마녀의 날)"""
    d = date(year, month, 1)
    days_until_friday = (4 - d.weekday()) % 7
    first_friday = d + timedelta(days=days_until_friday)
    return first_friday + timedelta(days=14)


def build_calendar_events(year, month):
    """주어진 년/월의 모든 캘린더 이벤트를 생성"""
    events = {}

    def add(date_str, category, title, detail=""):
        events.setdefault(date_str, []).append(
            {"category": category, "title": title, "detail": detail}
        )

    mi = month - 1  # 0-based

    # 1. 미국 경제지표
    for indicator, days in US_ECONOMIC_SCHEDULE.get(year, {}).items():
        if mi < len(days) and days[mi] > 0:
            add(f"{year}-{month:02d}-{days[mi]:02d}", "us_economic", indicator, f"{indicator} 발표")

    # 2. 중앙은행
    for m, d in FOMC_DATES.get(year, []):
        if m == month:
            add(f"{year}-{m:02d}-{d:02d}", "central_bank", "FOMC 금리결정", "미국 연준(Fed) 금리결정일")
    for m, d in ECB_DATES.get(year, []):
        if m == month:
            add(f"{year}-{m:02d}-{d:02d}", "central_bank", "ECB 금리결정", "유럽중앙은행 금리결정일")
    for m, d in BOJ_DATES.get(year, []):
        if m == month:
            add(f"{year}-{m:02d}-{d:02d}", "central_bank", "BOJ 금리결정", "일본은행 금리결정일")

    # 3. 한국 경제지표
    for m, d in BOK_DATES.get(year, []):
        if m == month:
            add(f"{year}-{m:02d}-{d:02d}", "kr_economic", "금통위 금리결정", "한국은행 금융통화위원회 기준금리 결정")
    for m, d in KR_CPI_DATES.get(year, []):
        if m == month:
            add(f"{year}-{m:02d}-{d:02d}", "kr_economic", "한국 CPI", "한국 소비자물가지수 발표")

    # 4. 네 마녀의 날 (3, 6, 9, 12월 셋째 금요일)
    if month in (3, 6, 9, 12):
        tf = get_third_friday(year, month)
        add(tf.isoformat(), "derivatives", "네 마녀의 날", "분기별 선물·옵션 동시 만기일 (Quadruple Witching)")

    # 5. 미국 실적발표 (yfinance)
    watchlist = load_cal_watchlist()
    for symbol in watchlist.get("us_earnings", []):
        try:
            t = yf.Ticker(symbol)
            cal = t.calendar
            if cal and isinstance(cal, dict):
                earnings = cal.get("Earnings Date")
                if earnings:
                    for ed in (earnings if isinstance(earnings, list) else [earnings]):
                        ds = ed.strftime("%Y-%m-%d") if hasattr(ed, "strftime") else str(ed)[:10]
                        try:
                            dd = date.fromisoformat(ds)
                            if dd.year == year and dd.month == month:
                                add(ds, "us_earnings", f"{symbol} 실적발표", f"{symbol} 분기 실적발표 예정")
                        except ValueError:
                            pass
        except Exception:
            pass

    # 6. 배당 이벤트 (yfinance)
    for symbol in watchlist.get("dividends", []):
        try:
            t = yf.Ticker(symbol)
            cal = t.calendar
            if cal and isinstance(cal, dict):
                ex_div = cal.get("Ex-Dividend Date")
                if ex_div:
                    ds = ex_div.strftime("%Y-%m-%d") if hasattr(ex_div, "strftime") else str(ex_div)[:10]
                    try:
                        dd = date.fromisoformat(ds)
                        if dd.year == year and dd.month == month:
                            add(ds, "dividend", f"{symbol} 배당락일", f"{symbol} 배당락일 (Ex-Dividend Date)")
                    except ValueError:
                        pass
                div_date = cal.get("Dividend Date")
                if div_date:
                    ds = div_date.strftime("%Y-%m-%d") if hasattr(div_date, "strftime") else str(div_date)[:10]
                    try:
                        dd = date.fromisoformat(ds)
                        if dd.year == year and dd.month == month:
                            add(ds, "dividend", f"{symbol} 배당지급일", f"{symbol} 배당지급일 (Dividend Date)")
                    except ValueError:
                        pass
        except Exception:
            pass

    # 7. 국내 실적/IR (DART API)
    for stock_code in watchlist.get("kr_earnings", []):
        for evt in fetch_dart_events(stock_code, year, month):
            add(evt["date"], "kr_earnings", evt["title"], evt["detail"])

    return events


@app.route("/api/calendar", methods=["GET"])
def get_calendar():
    """캘린더 이벤트 조회"""
    year = int(request.args.get("year", date.today().year))
    month = int(request.args.get("month", date.today().month))

    cache_key = f"{year}-{month}"
    now = time.time()
    if cache_key in _cal_cache and now - _cal_cache[cache_key][0] < _cal_cache_ttl:
        return jsonify(_cal_cache[cache_key][1])

    events = build_calendar_events(year, month)
    _cal_cache[cache_key] = (now, events)
    return jsonify(events)


@app.route("/api/calendar/watchlist", methods=["GET"])
def get_cal_watchlist_api():
    """캘린더 워치리스트 조회"""
    return jsonify(load_cal_watchlist())


@app.route("/api/calendar/watchlist", methods=["POST"])
def set_cal_watchlist_api():
    """캘린더 워치리스트 저장"""
    data = request.get_json()
    if not isinstance(data, dict):
        return jsonify({"error": "invalid format"}), 400
    save_cal_watchlist(data)
    return jsonify({"ok": True})


@app.route("/api/dart/company", methods=["GET"])
def dart_company_lookup():
    """DART 종목코드 → 회사명 조회"""
    code = request.args.get("code", "").replace(".KS", "").replace(".KQ", "")
    corp_map = load_dart_corp_map()
    info = corp_map.get(code)
    if info:
        return jsonify({"code": code, "name": info["name"]})
    return jsonify({"code": code, "name": ""}), 404


@app.route("/api/stock/search", methods=["GET"])
def stock_search():
    """종목 검색 (이름/코드) - MARKET_STOCKS + US 전체 + DART 전체"""
    q = request.args.get("q", "").strip().lower()
    if len(q) < 1:
        return jsonify([])

    results = []
    seen = set()

    # 1. MARKET_STOCKS (US, KR, JP, EU) - 우선 매칭
    for market, stocks in MARKET_STOCKS.items():
        for ticker, name, sector, industry in stocks:
            if q in ticker.lower() or q in name.lower():
                seen.add(ticker.lower().replace(".ks", "").replace(".kq", ""))
                results.append({"code": ticker, "name": name, "market": market})

    # 2. US 전체 종목 (나스닥 + NYSE)
    us_map = load_us_stock_map()
    us_extra = 0
    for symbol, info in us_map.items():
        if us_extra >= 15 or symbol.lower() in seen:
            continue
        if q in symbol.lower() or q in info["name"].lower():
            seen.add(symbol.lower())
            results.append({"code": symbol, "name": info["name"], "market": "us"})
            us_extra += 1

    # 3. DART corp map (추가 한국 종목)
    corp_map = load_dart_corp_map()
    kr_extra = 0
    for stock_code, info in corp_map.items():
        if kr_extra >= 10 or stock_code in seen:
            continue
        if q in stock_code or q in info["name"].lower():
            results.append({"code": stock_code, "name": info["name"], "market": "kr"})
            seen.add(stock_code)
            kr_extra += 1

    # 접두사 매칭 우선 정렬
    results.sort(key=lambda r: (
        0 if r["code"].lower().startswith(q) or r["name"].lower().startswith(q) else 1
    ))
    return jsonify(results[:15])


# ===== DART 공시 모니터링 =====

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")
TELEGRAM_TOKEN_GENERAL = os.environ.get("TELEGRAM_TOKEN_GENERAL")
TELEGRAM_TOKEN_EARNINGS = os.environ.get("TELEGRAM_TOKEN_EARNINGS")

TG_API_ID = os.environ.get("TG_API_ID")
TG_API_HASH = os.environ.get("TG_API_HASH")
TG_SESSION_PATH = os.environ.get("TG_SESSION_PATH", "sessions/tg_report")

_BD = os.path.dirname(os.path.abspath(__file__))
DART_MON_CFG_FILE = os.path.join(_BD, "dart_monitor_config.json")
DART_SEEN_FILE = os.path.join(_BD, "dart_seen.json")
DART_DAILY_FILE = os.path.join(_BD, "dart_daily.json")
DART_HISTORY_FILE = os.path.join(_BD, "dart_daily_history.json")
DART_SENT_FILE = os.path.join(_BD, "dart_sent.json")
EXCEL_PATH = os.path.expanduser(
    "~/Library/CloudStorage/GoogleDrive-changyun1222@gmail.com/내 드라이브/공시정리/DART_공시_누적.xlsx"
)
EXCEL_PATH_LOCAL = os.path.join(_BD, "DART_공시_누적.xlsx")

DEFAULT_DART_MON_CFG = {
    "watchlist": ["005930", "000660", "035420"],
    "earnings_conditions": {
        "revenue_growth": True, "op_profit_growth": True, "condition_type": "or",
    },
    "telegram_chat_ids": {"general": None, "earnings": None},
    "monitor_enabled": True,
    "digest_telegram_enabled": False,
}

_dart_alert_log = []
_monitor_thread = None


def _lj(path, default):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return default


def _sj(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_dm_cfg():
    cfg = _lj(DART_MON_CFG_FILE, None)
    if cfg is None:
        cfg = DEFAULT_DART_MON_CFG.copy()
        _sj(DART_MON_CFG_FILE, cfg)
    return cfg


def save_dm_cfg(cfg):
    _sj(DART_MON_CFG_FILE, cfg)


def _sync_dart_to_others(added_codes, removed_codes):
    try:
        cal = load_cal_watchlist()
        kr = cal.get("kr_earnings", [])
        for c in added_codes:
            if c not in kr:
                kr.append(c)
        for c in removed_codes:
            if c in kr:
                kr.remove(c)
        cal["kr_earnings"] = kr
        save_cal_watchlist(cal)
    except Exception as e:
        print(f"[cross-tab-sync] calendar sync failed: {e}", flush=True)
    try:
        corp_map = load_dart_corp_map()
        rcfg = telegram_report.get_config()
        rwl = rcfg.get("watchlist", [])
        for c in added_codes:
            info = corp_map.get(c)
            if info:
                name = info.get("name") if isinstance(info, dict) else info
                if name and name not in rwl:
                    rwl.append(name)
        for c in removed_codes:
            info = corp_map.get(c)
            if info:
                name = info.get("name") if isinstance(info, dict) else info
                if name and name in rwl:
                    rwl.remove(name)
        rcfg["watchlist"] = rwl
        telegram_report.save_config(rcfg)
    except Exception as e:
        print(f"[cross-tab-sync] report sync failed: {e}", flush=True)


def _migrate_cross_tab_sync():
    marker = os.path.join(_BD, "cache", "cross_tab_synced.flag")
    if os.path.exists(marker):
        return
    try:
        wl = load_dm_cfg().get("watchlist", [])
        if wl:
            _sync_dart_to_others(set(wl), set())
        with open(marker, "w") as f:
            pass
    except Exception as e:
        print(f"[cross-tab-sync] migration failed: {e}", flush=True)


def _load_seen():
    return set(_lj(DART_SEEN_FILE, []))


def _save_seen(seen):
    # 기존 순서를 유지하면서 최근 항목만 보관 (set은 순서 미보장이므로 리스트 기반)
    existing = _lj(DART_SEEN_FILE, [])
    existing_set = set(existing)
    new_items = [x for x in seen if x not in existing_set]
    ordered = existing + new_items
    _sj(DART_SEEN_FILE, ordered[-3000:])


def _load_daily_history():
    return _lj(DART_HISTORY_FILE, {"days": {}})


def _save_daily_history(h):
    _sj(DART_HISTORY_FILE, h)


def _load_daily():
    h = _load_daily_history()
    today = date.today().isoformat()
    items = h.get("days", {}).get(today, [])
    return {"date": today, "items": items}


def _save_daily(d):
    h = _load_daily_history()
    days = h.get("days", {})
    days[d["date"]] = d["items"]
    cutoff = (date.today() - timedelta(days=14)).isoformat()
    days = {k: v for k, v in days.items() if k >= cutoff}
    h["days"] = days
    _save_daily_history(h)


def _load_sent():
    return _lj(DART_SENT_FILE, {"date": "", "sent": []})


def _save_sent(d):
    _sj(DART_SENT_FILE, d)


def _alog(cat, msg):
    _dart_alert_log.append({"time": datetime.now().strftime("%H:%M"), "cat": cat, "msg": msg})
    if len(_dart_alert_log) > 50:
        _dart_alert_log.pop(0)


# --- Telegram ---
TG_MAX_LEN = 4000


def _tg_split_message(text, max_len=TG_MAX_LEN):
    if len(text) <= max_len:
        return [text]
    chunks = []
    current = ""
    for para in text.split("\n\n"):
        if len(para) > max_len:
            if current:
                chunks.append(current)
                current = ""
            for line in para.split("\n"):
                if len(line) > max_len:
                    while line:
                        space = max_len if not current else max_len - len(current) - 1
                        if current:
                            current += "\n" + line[:space]
                            chunks.append(current)
                            current = ""
                        else:
                            chunks.append(line[:space])
                        line = line[space:]
                elif current and len(current) + 1 + len(line) > max_len:
                    chunks.append(current)
                    current = line
                else:
                    current = current + "\n" + line if current else line
        elif current and len(current) + 2 + len(para) > max_len:
            chunks.append(current)
            current = para
        else:
            current = current + "\n\n" + para if current else para
    if current:
        chunks.append(current)
    return chunks


def _tg_send(token, chat_id, text):
    if not chat_id:
        return False
    try:
        chunks = _tg_split_message(text)
        total = len(chunks)
        if total > 1:
            print(f"[TG] 메시지 분할 전송: {total}개 청크")
        for idx, chunk in enumerate(chunks, 1):
            if total > 1:
                chunk += f"\n\n({idx}/{total})"
            requests.post(
                f"https://api.telegram.org/bot{token}/sendMessage",
                json={"chat_id": chat_id, "text": chunk,
                      "parse_mode": "HTML", "disable_web_page_preview": True},
                timeout=10,
            )
            if idx < total:
                time.sleep(0.3)
        return True
    except Exception:
        return False


def _tg_detect(token):
    try:
        res = requests.get(f"https://api.telegram.org/bot{token}/getUpdates", timeout=10)
        data = res.json()
        if data.get("ok") and data.get("result"):
            return data["result"][-1]["message"]["chat"]["id"]
    except Exception:
        pass
    return None


# --- Gemini ---
def _gemini(prompt):
    try:
        res = requests.post(
            f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash-lite:generateContent"
            f"?key={GEMINI_API_KEY}",
            json={"contents": [{"parts": [{"text": prompt}]}]},
            timeout=30,
        )
        return res.json()["candidates"][0]["content"]["parts"][0]["text"]
    except Exception:
        return ""


# --- DART 문서 텍스트 추출 ---
def _dart_doc(rcept_no):
    try:
        res = requests.get(
            "https://opendart.fss.or.kr/api/document.xml",
            params={"crtfc_key": DART_API_KEY, "rcept_no": rcept_no},
            timeout=30,
        )
        with zipfile.ZipFile(io.BytesIO(res.content)) as z:
            for name in z.namelist():
                if name.endswith(".xml"):
                    with z.open(name) as f:
                        raw = f.read().decode("utf-8", errors="ignore")
                        text = re.sub(r"<[^>]+>", " ", raw)
                        return re.sub(r"\s+", " ", text).strip()[:8000]
    except Exception:
        pass
    return ""


def _mcap_str(stock_code):
    try:
        tc = stock_code + ".KS" if stock_code and "." not in stock_code else stock_code
        mc = yf.Ticker(tc).info.get("marketCap", 0)
        if mc >= 1e12:
            return f"{mc / 1e12:.1f}조원"
        if mc >= 1e8:
            return f"{mc / 1e8:.0f}억원"
    except Exception:
        pass
    return "-"


def _fv(v):
    return f"{v:,.0f}억" if v is not None else "-"


# --- Feature 1: 종목 모니터링 ---
def _ck_watchlist():
    cfg = load_dm_cfg()
    seen = _load_seen()
    cmap = load_dart_corp_map()
    td = date.today().strftime("%Y%m%d")
    cid = cfg["telegram_chat_ids"].get("general")
    if not cid:
        return

    chg = False
    for sc in cfg.get("watchlist", []):
        info = cmap.get(sc)
        if not info:
            continue
        try:
            page_no = 1
            while True:
                res = requests.get(
                    "https://opendart.fss.or.kr/api/list.json",
                    params={"crtfc_key": DART_API_KEY, "corp_code": info["corp_code"],
                            "bgn_de": td, "end_de": td, "page_no": page_no, "page_count": 100},
                    timeout=10,
                )
                data = res.json()
                if data.get("status") != "000":
                    break
                items = data.get("list", [])
                if not items:
                    break
                for it in items:
                    rno = it.get("rcept_no", "")
                    if rno in seen:
                        continue
                    seen.add(rno)
                    chg = True
                    title = it.get("report_nm", "")
                    doc = _dart_doc(rno)
                    summary = _gemini(
                        f"다음 DART 공시의 핵심 내용을 3줄 이내로 간결하게 한국어로 요약해줘:\n\n{doc[:4000]}"
                    ) if doc else ""
                    msg = (f"📢 <b>{info['name']}</b> 새 공시\n\n"
                           f"📋 {title}\n📅 {it.get('rcept_dt', '')}\n")
                    if summary:
                        msg += f"\n📝 {summary}\n"
                    msg += f"\n🔗 <a href='https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rno}'>공시 원문</a>"
                    _tg_send(TELEGRAM_TOKEN_GENERAL, cid, msg)
                    _alog("종목", f"{info['name']} - {title}")
                total_page = int(data.get("total_page", 1))
                if page_no >= total_page:
                    break
                page_no += 1
                time.sleep(0.2)
        except Exception:
            continue
    if chg:
        _save_seen(seen)


def _fmt_amt(v):
    if v is None:
        return "-"
    try:
        n = float(v)
    except (TypeError, ValueError):
        return str(v)
    if abs(n) >= 10000:
        return f"{n/10000:.1f}조"
    return f"{n:,.0f}억"


def _pct_change(cur, prev):
    try:
        c, p = float(cur), float(prev)
        if p == 0:
            return ""
        pct = (c - p) / abs(p) * 100
        return f"{pct:+.0f}%"
    except (TypeError, ValueError):
        return ""


def _majorstock_summary(corp_code, rno):
    try:
        res = requests.get(
            "https://opendart.fss.or.kr/api/majorstock.json",
            params={"crtfc_key": DART_API_KEY, "corp_code": corp_code},
            timeout=15,
        )
        data = res.json()
        if data.get("status") != "000":
            return ""
        hit = next((x for x in data.get("list", []) if x.get("rcept_no") == rno), None)
        if not hit:
            return ""
        repror = hit.get("repror", "")
        report_tp = hit.get("report_tp", "")
        try:
            curr = float(hit.get("stkrt", "0"))
            irds = float(hit.get("stkrt_irds", "0"))
            prev = round(curr - irds, 2)
        except (TypeError, ValueError):
            return ""
        reason = (hit.get("report_resn") or "").lstrip("- ").strip()
        if len(reason) > 50:
            reason = reason[:50]
        parts = []
        if repror:
            parts.append(f"보고자: {repror}")
        parts.append(f"보유비율: {prev}%→{curr}%")
        if reason:
            parts.append(f"사유: {reason}")
        if report_tp:
            parts.append(f"({report_tp})")
        return " · ".join(parts)
    except Exception:
        return ""


def _dart_summary(title, rno, corp_code=""):
    try:
        doc = _dart_doc(rno)
    except Exception:
        doc = ""

    if not doc and "대량보유" in title and corp_code:
        return _majorstock_summary(corp_code, rno)

    if not doc:
        return ""

    if "공급계약" in title:
        raw = _gemini(
            "다음 DART 공급계약 공시에서 정보를 추출해. 반드시 JSON만 응답:\n"
            '{"counterparty":"계약상대방","amount":계약금액(억원),'
            '"recent_revenue":최근매출액(억원),"pct":매출대비퍼센트숫자,'
            '"period":"계약기간 문자열"}\n'
            f"숫자는 억원 단위, 없으면 null.\n\n{doc[:4000]}"
        )
        try:
            m = re.search(r"\{[\s\S]*\}", raw)
            if m:
                d = json.loads(m.group())
                parts = []
                if d.get("counterparty"):
                    parts.append(f"계약상대: {d['counterparty']}")
                if d.get("amount") is not None:
                    parts.append(f"금액: {_fmt_amt(d['amount'])}")
                if d.get("recent_revenue") is not None:
                    parts.append(f"최근매출: {_fmt_amt(d['recent_revenue'])}")
                if d.get("pct") is not None:
                    parts.append(f"매출대비: {d['pct']}%")
                if d.get("period"):
                    parts.append(f"기간: {d['period']}")
                if parts:
                    return " · ".join(parts)
        except Exception:
            pass
        return _gemini(f"다음 DART 공시의 핵심 내용을 1줄로 요약:\n{doc[:3000]}") or ""

    if "대량보유" in title:
        raw = _gemini(
            "다음 DART 대량보유 공시에서 정보를 추출해. 반드시 JSON만 응답:\n"
            '{"reporter":"보고자","prev_ratio":직전보유비율숫자,'
            '"curr_ratio":이번보유비율숫자,"reason":"보고사유",'
            '"purpose":"보유목적"}\n'
            f"숫자는 퍼센트, 없으면 null.\n\n{doc[:4000]}"
        )
        try:
            m = re.search(r"\{[\s\S]*\}", raw)
            if m:
                d = json.loads(m.group())
                parts = []
                if d.get("reporter"):
                    parts.append(f"보고자: {d['reporter']}")
                pr = d.get("prev_ratio")
                cr = d.get("curr_ratio")
                if pr is not None and cr is not None:
                    parts.append(f"보유비율: {pr}%→{cr}%")
                elif cr is not None:
                    parts.append(f"보유비율: {cr}%")
                if d.get("reason"):
                    parts.append(f"사유: {d['reason']}")
                if d.get("purpose"):
                    parts.append(f"목적: {d['purpose']}")
                if parts:
                    return " · ".join(parts)
        except Exception:
            pass
        return _gemini(f"다음 DART 공시의 핵심 내용을 1줄로 요약:\n{doc[:3000]}") or ""

    if "잠정실적" in title or "영업(잠정)" in title:
        raw = _gemini(
            "다음 DART 잠정실적 공시에서 정보를 추출해. 반드시 JSON만 응답:\n"
            '{"revenue":{"current":당기매출(억원),"prev":전기매출(억원)},'
            '"op_profit":{"current":당기영업이익(억원),"prev":전기영업이익(억원)}}\n'
            f"숫자는 억원 단위, 없으면 null.\n\n{doc[:4000]}"
        )
        try:
            m = re.search(r"\{[\s\S]*\}", raw)
            if m:
                d = json.loads(m.group())
                rv = d.get("revenue", {})
                op = d.get("op_profit", {})
                parts = []
                rc, rp = rv.get("current"), rv.get("prev")
                oc, op_ = op.get("current"), op.get("prev")
                if rc is not None:
                    s = f"매출: {_fmt_amt(rc)}"
                    if rp is not None:
                        s += f"(전기 {_fmt_amt(rp)}, {_pct_change(rc, rp)})"
                    parts.append(s)
                if oc is not None:
                    s = f"영업익: {_fmt_amt(oc)}"
                    if op_ is not None:
                        s += f"(전기 {_fmt_amt(op_)}, {_pct_change(oc, op_)})"
                    parts.append(s)
                if parts:
                    return " · ".join(parts)
        except Exception:
            pass
        return _gemini(f"다음 DART 공시의 핵심 내용을 1줄로 요약:\n{doc[:3000]}") or ""

    return _gemini(f"다음 DART 공시의 핵심 내용을 1줄로 요약:\n{doc[:3000]}") or ""


# --- Feature 2: 잠정실적 ---
def _ck_earnings():
    cfg = load_dm_cfg()
    seen = _load_seen()
    td = date.today().strftime("%Y%m%d")
    cid = cfg["telegram_chat_ids"].get("earnings")
    if not cid:
        return
    conds = cfg.get("earnings_conditions", {})

    chg = False
    try:
        page_no = 1
        while True:
            res = requests.get(
                "https://opendart.fss.or.kr/api/list.json",
                params={"crtfc_key": DART_API_KEY, "bgn_de": td, "end_de": td,
                        "pblntf_ty": "B", "page_no": page_no, "page_count": 100},
                timeout=10,
            )
            data = res.json()
            if data.get("status") != "000":
                break
            items = data.get("list", [])
            if not items:
                break
            for it in items:
                title = it.get("report_nm", "")
                if "잠정실적" not in title and "영업(잠정)" not in title:
                    continue
                rno = it.get("rcept_no", "")
                key = f"e_{rno}"
                if key in seen:
                    continue
                seen.add(key)
                chg = True

                corp_name = it.get("corp_name", "")
                stock_code = it.get("stock_code", "")

                doc = _dart_doc(rno)
                extract = _gemini(
                    "다음 DART 잠정실적 공시에서 정보를 추출해. 반드시 JSON만 응답:\n"
                    '{"fs_type":"연결 또는 별도",'
                    '"revenue":{"current":당기매출(억원),"prev":전기매출(억원)},'
                    '"op_profit":{"current":당기영업이익(억원),"prev":전기영업이익(억원)},'
                    '"net_income":{"current":당기순이익(억원),"prev":전기순이익(억원)},'
                    '"reason":"변동요인 2줄 요약"}\n'
                    f"숫자는 억원 단위, 없으면 null.\n\n{doc[:5000]}"
                )

                fd = None
                try:
                    m = re.search(r"\{[\s\S]*\}", extract)
                    if m:
                        fd = json.loads(m.group())
                except Exception:
                    pass

                if fd:
                    rev = fd.get("revenue", {})
                    op = fd.get("op_profit", {})
                    rg = (rev.get("current") or 0) > (rev.get("prev") or 0) if rev.get("current") and rev.get("prev") else False
                    og = (op.get("current") or 0) > (op.get("prev") or 0) if op.get("current") and op.get("prev") else False
                    ct = conds.get("condition_type", "or")
                    cr = conds.get("revenue_growth", True)
                    co = conds.get("op_profit_growth", True)
                    passed = ((cr and rg) or (co and og)) if ct == "or" else ((not cr or rg) and (not co or og))
                    if not passed:
                        continue

                mc = _mcap_str(stock_code)
                msg = "📊 <b>잠정실적 공시 알림</b>\n\n"
                msg += f"📅 공시일시: {it.get('rcept_dt', '')}\n"
                msg += f"🏢 기업명: <b>{corp_name}</b> (시총: {mc})\n"
                msg += f"📋 보고서: {title}\n"
                if fd:
                    msg += f"📑 재무제표: {fd.get('fs_type', '-')}\n\n"
                    r_ = fd.get("revenue", {})
                    o_ = fd.get("op_profit", {})
                    n_ = fd.get("net_income", {})
                    msg += f"💰 매출액: {_fv(r_.get('current'))} (전기: {_fv(r_.get('prev'))})\n"
                    msg += f"💰 영업이익: {_fv(o_.get('current'))} (전기: {_fv(o_.get('prev'))})\n"
                    msg += f"💰 순이익: {_fv(n_.get('current'))} (전기: {_fv(n_.get('prev'))})\n"
                    reason = fd.get("reason", "")
                    if reason:
                        msg += f"\n📝 변동요인: {reason}\n"
                msg += f"\n🔗 <a href='https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rno}'>DART 공시 원문</a>"
                if stock_code:
                    msg += f"\n🔗 <a href='https://finance.naver.com/item/main.naver?code={stock_code}'>네이버 회사정보</a>"
                _tg_send(TELEGRAM_TOKEN_EARNINGS, cid, msg)
                _alog("잠정", f"{corp_name} - {title}")
            total_page = int(data.get("total_page", 1))
            if page_no >= total_page:
                break
            page_no += 1
            time.sleep(0.2)
    except Exception:
        pass
    if chg:
        _save_seen(seen)


# --- Feature 3: 주요 공시 ---
_MAJOR_KW = ["대량보유", "실적", "매출액", "공급계약"]
_MAJOR_EXCLUDE = ["증권발행실적보고서", "주식등의대량보유상황보고서(일반)"]


def _ck_major():
    seen = _load_seen()
    daily = _load_daily()
    td = date.today().strftime("%Y%m%d")
    ti = date.today().isoformat()
    if daily.get("date") != ti:
        daily = {"date": ti, "items": []}

    chg = False
    try:
        page_no = 1
        while True:
            res = requests.get(
                "https://opendart.fss.or.kr/api/list.json",
                params={"crtfc_key": DART_API_KEY, "bgn_de": td, "end_de": td,
                        "page_no": page_no, "page_count": 100},
                timeout=10,
            )
            data = res.json()
            if data.get("status") != "000":
                break
            items = data.get("list", [])
            if not items:
                break
            for it in items:
                title = it.get("report_nm", "")
                if not any(k in title for k in _MAJOR_KW):
                    continue
                if any(ex in title for ex in _MAJOR_EXCLUDE):
                    continue
                rno = it.get("rcept_no", "")
                key = f"m_{rno}"
                if key in seen:
                    continue
                seen.add(key)
                chg = True
                corp_name = it.get("corp_name", "")
                stock_code = it.get("stock_code", "")
                it_corp_code = it.get("corp_code", "")
                if not it_corp_code and stock_code:
                    cmap = load_dart_corp_map()
                    entry = cmap.get(stock_code, {})
                    it_corp_code = entry.get("corp_code", "") if isinstance(entry, dict) else str(entry)
                summary = _dart_summary(title, rno, corp_code=it_corp_code)
                daily["items"].append({
                    "corp_name": corp_name,
                    "stock_code": stock_code,
                    "corp_code": it_corp_code,
                    "title": title, "rcept_no": rno,
                    "rcept_dt": it.get("rcept_dt", ""),
                    "detected_at": datetime.now().strftime("%H:%M"),
                    "summary": summary,
                    "summary_tries": 1 if not summary else 0,
                })
                _alog("주요", f"{corp_name} - {title}")
            total_page = int(data.get("total_page", 1))
            if page_no >= total_page:
                break
            page_no += 1
            time.sleep(0.2)
    except Exception:
        pass

    retried = 0
    for item in daily.get("items", []):
        if retried >= 12:
            break
        if item.get("summary"):
            continue
        tries = item.get("summary_tries", 0)
        if tries >= 5:
            continue
        cc = item.get("corp_code", "")
        if not cc and item.get("stock_code"):
            cmap = load_dart_corp_map()
            entry = cmap.get(item["stock_code"], {})
            cc = entry.get("corp_code", "") if isinstance(entry, dict) else str(entry)
            item["corp_code"] = cc
        try:
            s = _dart_summary(item.get("title", ""), item.get("rcept_no", ""), corp_code=cc)
            item["summary_tries"] = tries + 1
            if s:
                item["summary"] = s
                chg = True
        except Exception:
            item["summary_tries"] = tries + 1
        retried += 1
        time.sleep(0.5)

    if retried:
        chg = True

    if chg:
        _save_seen(seen)
        _save_daily(daily)


_DAILY_SLOTS = [
    {"hour": 11, "key": "11", "start": "00:00", "end": "10:59",
     "header": "📋 <b>오전 공시 정리 (00:00~11:00)</b>"},
    {"hour": 14, "key": "14", "start": "11:00", "end": "13:59",
     "header": "📋 <b>오후 공시 정리 (11:00~14:00)</b>"},
    {"hour": 18, "key": "18", "start": "14:00", "end": "17:59",
     "header": "📋 <b>오후 공시 정리 (14:00~18:00)</b>"},
    {"hour": 22, "key": "22", "start": None, "end": None,
     "header": None},
]


def _send_daily_slot(slot):
    cfg = load_dm_cfg()
    daily = _load_daily()
    cid = cfg["telegram_chat_ids"].get("general")
    if not cid or not daily.get("items"):
        return
    if slot["start"] is not None:
        filtered = [it for it in daily["items"]
                    if slot["start"] <= it.get("detected_at", "00:00") <= slot["end"]]
        if not filtered:
            return
        header = slot["header"]
    else:
        filtered = daily["items"]
        if not filtered:
            return
        header = f"📋 <b>오늘의 전체 공시 정리 ({date.today().isoformat()})</b>"
    msg = f"{header}\n\n"
    for i, it in enumerate(filtered, 1):
        msg += f"{i}. <b>{it['corp_name']}</b>\n   📋 {it['title']}\n"
        msg += f"   🔗 https://dart.fss.or.kr/dsaf001/main.do?rcpNo={it['rcept_no']}\n\n"
    _tg_send(TELEGRAM_TOKEN_GENERAL, cid, msg)
    _alog("일일", f"{slot['key']}시 공시 {len(filtered)}건 전송")


def _save_excel():
    import shutil
    from openpyxl import Workbook, load_workbook
    daily = _load_daily()
    if not daily.get("items"):
        return

    # 항상 로컬에 저장
    try:
        wb = load_workbook(EXCEL_PATH_LOCAL)
        ws = wb.active
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.title = "DART 공시"
        ws.append(["날짜", "종목명", "공시사유", "섹터", "시총", "매출액", "영업이익", "공시내용", "영익대비시총"])

    for it in daily["items"]:
        sc = it.get("stock_code", "")
        mc_s = rev_s = op_s = sector = ""
        try:
            tc = sc + ".KS" if sc and "." not in sc else sc
            if tc:
                inf = yf.Ticker(tc).info
                mc = inf.get("marketCap", 0)
                mc_s = f"{mc / 1e8:.0f}억" if mc else ""
                rv = inf.get("totalRevenue", 0)
                rev_s = f"{rv / 1e8:.0f}억" if rv else ""
                oi = inf.get("operatingIncome") or inf.get("ebitda", 0)
                op_s = f"{oi / 1e8:.0f}억" if oi else ""
                sector = inf.get("sector", "")
        except Exception:
            pass
        if not sector:
            for s in MARKET_STOCKS.get("kr", []):
                if sc and sc in s[0]:
                    sector = s[2]
                    break
        doc = _dart_doc(it.get("rcept_no", ""))
        summary = _gemini(f"다음 DART 공시의 핵심 내용을 1줄로 요약:\n{doc[:3000]}") if doc else ""
        ratio = ""
        try:
            if mc_s and op_s:
                mcv = float(mc_s.replace("억", "").replace(",", ""))
                opv = float(op_s.replace("억", "").replace(",", ""))
                if opv > 0:
                    ratio = f"{mcv / opv:.1f}배"
        except Exception:
            pass
        ws.append([it.get("rcept_dt", ""), it.get("corp_name", ""), it.get("title", ""),
                   sector, mc_s, rev_s, op_s, summary[:200], ratio])

    try:
        wb.save(EXCEL_PATH_LOCAL)
        _alog("엑셀", f"공시 {len(daily['items'])}건 로컬 저장 완료")
    except Exception as e:
        _alog("엑셀", f"로컬 저장 실패: {e}")
        return

    # Google Drive로 복사
    try:
        os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
        shutil.copy2(EXCEL_PATH_LOCAL, EXCEL_PATH)
        _alog("엑셀", "Google Drive 복사 완료")
    except Exception as e:
        _alog("엑셀", f"Google Drive 복사 실패 (로컬 저장은 완료): {e}")


# --- Background monitor ---
def _dart_loop():
    time.sleep(10)
    while True:
        if datetime.now().weekday() >= 5:
            time.sleep(300)
            continue

        try:
            cfg = load_dm_cfg()
            if cfg.get("monitor_enabled", True):
                _ck_watchlist()
                _ck_earnings()
                _ck_major()
                now = datetime.now()
                today_iso = now.date().isoformat()
                sent_state = _load_sent()
                if sent_state.get("date") != today_iso:
                    sent_state = {"date": today_iso, "sent": []}
                cfg = load_dm_cfg()
                tg_on = cfg.get("digest_telegram_enabled", False)
                for slot in _DAILY_SLOTS:
                    if now.hour == slot["hour"] and slot["key"] not in sent_state["sent"]:
                        if tg_on:
                            _send_daily_slot(slot)
                        sent_state["sent"].append(slot["key"])
                        _save_sent(sent_state)
                        if slot["key"] == "22":
                            _save_excel()
        except Exception as e:
            print(f"[DART Monitor] {e}")
        time.sleep(300)


def _start_monitor():
    global _monitor_thread
    if _monitor_thread and _monitor_thread.is_alive():
        return
    _monitor_thread = threading.Thread(target=_dart_loop, daemon=True)
    _monitor_thread.start()


_start_monitor()


# --- Earnings tracker (잠정실적 추적) ---
def _earnings_tracker_loop():
    """매일 평일 20:00 KST에 잠정실적 트래커 실행. 시작 시 catch-up 포함."""
    import earnings_tracker

    try:
        state = earnings_tracker.load_state()
        today = date.today()
        last_run_str = state.get("last_run_date")
        if today.weekday() < 5 and last_run_str != today.isoformat():
            print(f"[earnings_tracker] 시작 시 catch-up: last_run={last_run_str}, today={today.isoformat()}", flush=True)
            earnings_tracker.run_daily()
    except Exception as e:
        import traceback
        print(f"[earnings_tracker_loop] catch-up 에러: {e}", flush=True)
        traceback.print_exc()

    while True:
        try:
            now = datetime.now()
            target = now.replace(hour=20, minute=0, second=0, microsecond=0)
            if now >= target:
                target = target + timedelta(days=1)

            sleep_sec = (target - now).total_seconds()
            print(f"[earnings_tracker] 다음 실행까지 {int(sleep_sec)}초 대기 (target: {target})", flush=True)
            time.sleep(sleep_sec)

            if datetime.now().weekday() >= 5:
                continue

            earnings_tracker.run_daily()

        except Exception as e:
            import traceback
            print(f"[earnings_tracker_loop] 예상치 못한 에러: {e}", flush=True)
            traceback.print_exc()
            time.sleep(3600)


threading.Thread(target=_earnings_tracker_loop, daemon=True).start()

try:
    _migrate_cross_tab_sync()
except Exception as e:
    print("cross-tab sync migration failed:", e, flush=True)


@app.route("/api/earnings-tracker/run", methods=["POST"])
def trigger_earnings_tracker():
    if not _is_private_ip(request.remote_addr):
        return jsonify({"error": "forbidden"}), 403
    import earnings_tracker

    backfill_days = request.args.get("backfill_days", type=int)
    if backfill_days is not None:
        results = earnings_tracker.run_backfill(backfill_days)
        return jsonify({"mode": "backfill", "results": results})

    force = request.args.get("force", "false").lower() == "true"
    result = earnings_tracker.run_daily(force=force)
    return jsonify(result)


# --- DART Daily API ---
@app.route("/api/dart/daily", methods=["GET"])
def get_dart_daily():
    days_param = request.args.get("days")
    if days_param:
        try:
            n = int(days_param)
        except ValueError:
            n = 1
        h = _load_daily_history()
        all_days = h.get("days", {})
        cutoff = (date.today() - timedelta(days=n - 1)).isoformat()
        items = []
        for d, d_items in all_days.items():
            if d >= cutoff:
                for it in d_items:
                    it_copy = dict(it)
                    it_copy["date"] = d
                    items.append(it_copy)
        return jsonify({"days": n, "items": items})
    dt = request.args.get("date", date.today().isoformat())
    h = _load_daily_history()
    items = h.get("days", {}).get(dt, [])
    return jsonify({"date": dt, "items": items})


@app.route("/api/dart/dates", methods=["GET"])
def get_dart_dates():
    h = _load_daily_history()
    dates = sorted(h.get("days", {}).keys(), reverse=True)
    return jsonify({"dates": dates})


# --- DART Monitor API ---
@app.route("/api/dart/monitor/config", methods=["GET"])
def get_dm_cfg():
    return jsonify(load_dm_cfg())


@app.route("/api/dart/monitor/config", methods=["POST"])
def set_dm_cfg():
    old_wl = set(load_dm_cfg().get("watchlist", []))
    new_cfg = request.get_json()
    new_wl = set(new_cfg.get("watchlist", []))
    save_dm_cfg(new_cfg)
    added = new_wl - old_wl
    removed = old_wl - new_wl
    if added or removed:
        _sync_dart_to_others(added, removed)
    return jsonify({"ok": True})


@app.route("/api/dart/monitor/status", methods=["GET"])
def get_dm_status():
    cfg = load_dm_cfg()
    return jsonify({
        "running": _monitor_thread is not None and _monitor_thread.is_alive(),
        "enabled": cfg.get("monitor_enabled", True),
    })


@app.route("/api/dart/monitor/log", methods=["GET"])
def get_dm_log():
    return jsonify(_dart_alert_log[-20:])


@app.route("/api/dart/telegram/test", methods=["POST"])
def test_dm_tg():
    body = request.get_json()
    tt = body.get("type", "general")
    token = TELEGRAM_TOKEN_GENERAL if tt == "general" else TELEGRAM_TOKEN_EARNINGS
    cid = _tg_detect(token)
    if not cid:
        return jsonify({"ok": False, "error": "봇에 먼저 /start 메시지를 보내주세요"})
    cfg = load_dm_cfg()
    cfg["telegram_chat_ids"][tt] = cid
    save_dm_cfg(cfg)
    ok = _tg_send(token, cid, "✅ 텔레그램 연결 성공! DART 공시 알림이 활성화됩니다.")
    return jsonify({"ok": ok, "chat_id": cid})


@app.route("/api/dart/monitor/toggle", methods=["POST"])
def toggle_dm():
    cfg = load_dm_cfg()
    cfg["monitor_enabled"] = not cfg.get("monitor_enabled", True)
    save_dm_cfg(cfg)
    if cfg["monitor_enabled"]:
        _start_monitor()
    return jsonify({"ok": True, "enabled": cfg["monitor_enabled"]})


@app.route("/api/dart/daily/trigger", methods=["POST"])
def trigger_daily():
    """수동으로 전체 공시 정리 + 엑셀 저장을 실행"""
    try:
        full_slot = _DAILY_SLOTS[-1]
        _send_daily_slot(full_slot)
        _save_excel()
        return jsonify({"ok": True, "msg": "전체 공시 정리 및 엑셀 저장 완료"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ===== 증권사 Report =====

@app.route("/api/report/channels", methods=["GET"])
def get_report_channels():
    return jsonify(telegram_report.get_config()["channels"])


@app.route("/api/report/channels", methods=["POST"])
def update_report_channels():
    body = request.json
    action = body.get("action")
    channel = body.get("channel", "").strip()
    cfg = telegram_report.get_config()
    channels = cfg.get("channels", [])
    if action == "add" and channel and channel not in channels:
        channels.append(channel)
    elif action == "remove" and channel in channels:
        channels.remove(channel)
    cfg["channels"] = channels
    telegram_report.save_config(cfg)
    return jsonify(channels)


@app.route("/api/report/search", methods=["POST"])
def start_report_search():
    body = request.json
    channel = body.get("channel", "")
    keyword = body.get("keyword", "")
    date_from = datetime.strptime(body.get("date_from", "2020-01-01"), "%Y-%m-%d")
    date_to = datetime.strptime(body.get("date_to", datetime.now().strftime("%Y-%m-%d")), "%Y-%m-%d")
    cfg = telegram_report.get_config()
    download_base = cfg.get("download_base", telegram_report.DEFAULT_CONFIG["download_base"])
    job_id = telegram_report.start_search_job(
        channel, keyword, date_from, date_to,
        TG_API_ID, TG_API_HASH, TG_SESSION_PATH, download_base,
    )
    return jsonify({"job_id": job_id})


@app.route("/api/report/status/<job_id>", methods=["GET"])
def get_report_status(job_id):
    job = telegram_report.get_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify(job)


@app.route("/api/report/stop/<job_id>", methods=["POST"])
def stop_report_search(job_id):
    job = telegram_report.get_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    job["stop_requested"] = True
    return jsonify({"ok": True})


@app.route("/api/dart-report/download", methods=["POST"])
def start_dart_download():
    body = request.json or {}
    code = body.get("code", "")
    name = body.get("name", "")
    date_from = body.get("date_from", "")
    date_to = body.get("date_to", "")
    want_xml = body.get("want_xml", True)
    want_pdf = body.get("want_pdf", True)
    corp_code = earnings_tracker.get_corp_code(code)
    if not corp_code:
        return jsonify({"error": "corp_code를 찾을 수 없습니다"}), 400
    cfg = telegram_report.get_config()
    download_base = cfg.get("download_base", telegram_report.DEFAULT_CONFIG["download_base"])
    job_id = dart_report.start_download_job(
        corp_code, name, date_from, date_to, download_base, want_xml, want_pdf,
    )
    return jsonify({"job_id": job_id})


@app.route("/api/dart-report/status/<job_id>", methods=["GET"])
def get_dart_report_status(job_id):
    job = dart_report.get_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify(job)


@app.route("/api/dart-report/stop/<job_id>", methods=["POST"])
def stop_dart_download(job_id):
    job = dart_report.get_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    job["stop_requested"] = True
    return jsonify({"ok": True})


@app.route("/api/report/config", methods=["GET"])
def get_report_config():
    return jsonify(telegram_report.get_config())


@app.route("/api/report/index-status", methods=["GET"])
def get_report_index_status():
    cfg = telegram_report.get_config()
    channels = cfg.get("channels", [])
    if not channels:
        return jsonify([])
    result = []
    for ch in channels:
        result.append(telegram_report.get_index_status(ch))
    return jsonify(result)


@app.route("/api/report/watchlist", methods=["POST"])
def update_report_watchlist():
    body = request.json
    action = body.get("action")
    stock = body.get("stock", "").strip()
    cfg = telegram_report.get_config()
    wl = cfg.get("watchlist", [])
    if action == "add" and stock and stock not in wl:
        wl.append(stock)
    elif action == "remove" and stock in wl:
        wl.remove(stock)
    cfg["watchlist"] = wl
    telegram_report.save_config(cfg)
    return jsonify(wl)


@app.route("/api/report/settings", methods=["GET"])
def get_report_settings():
    cfg = telegram_report.get_config()
    return jsonify({
        "personal_channels": cfg.get("personal_channels", []),
        "forward_enabled": cfg.get("forward_enabled", True),
    })


@app.route("/api/report/settings", methods=["POST"])
def update_report_settings():
    body = request.json
    action = body.get("action")
    channel = body.get("channel", "").strip()
    cfg = telegram_report.get_config()
    channels = cfg.get("personal_channels", [])
    if action == "add" and channel and channel not in channels:
        channels.append(channel)
    elif action == "remove" and channel in channels:
        channels.remove(channel)
    cfg["personal_channels"] = channels
    telegram_report.save_config(cfg)
    return jsonify(channels)


if __name__ == "__main__":
    app.run(host='0.0.0.0', port=8000, debug=False)

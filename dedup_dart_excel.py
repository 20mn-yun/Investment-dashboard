import os
import shutil
from datetime import datetime
from openpyxl import load_workbook, Workbook

PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "DART_공시_누적.xlsx")

if not os.path.exists(PATH):
    print("파일 없음:", PATH)
    raise SystemExit

wb = load_workbook(PATH)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))

if not rows:
    print("빈 파일입니다.")
    raise SystemExit

header = rows[0]
data = rows[1:]

backup = PATH + ".bak_" + datetime.now().strftime("%Y%m%d_%H%M%S")
shutil.copy2(PATH, backup)
print("백업 생성:", backup)

seen = set()
kept = []
removed = []
for r in data:
    if len(r) < 3:
        kept.append(r)
        continue
    key = (r[0], r[1], r[2])
    if key in seen:
        removed.append(r)
        continue
    seen.add(key)
    kept.append(r)

print("원본 데이터 행:", len(data))
print("중복 제거 후:", len(kept))
print("삭제된 행:", len(removed))
print("--- 삭제된 항목 미리보기 (최대 30개) ---")
for r in removed[:30]:
    print(r[0], "|", r[1], "|", r[2])

nwb = Workbook()
nws = nwb.active
nws.title = ws.title
nws.append(list(header))
for r in kept:
    nws.append(list(r))
nwb.save(PATH)
print("정리 완료:", PATH)

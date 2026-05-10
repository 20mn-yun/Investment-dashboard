"""
DART 영업(잠정)실적 공시 추적기.

매일 평일 20:00 KST 실행. DART에서 당일 '연결재무제표 기준 영업(잠정)실적' 공시를
감지해 Google Drive 엑셀(잠정실적_누적.xlsx)에 누적 기록.

Phase 1: 메타데이터(기업명, 공시일, 종목코드, DART 링크)만 기록.
Phase 2~4에서 재무 수치 파싱, 과거 분기, 필터링 추가 예정.
"""

import io
import os
import re
import json
import zipfile
import xml.etree.ElementTree as ET
import requests
import anthropic
from datetime import datetime, date, timedelta
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

load_dotenv()

DART_API_KEY = os.environ.get("DART_API_KEY")
EXCEL_PATH = os.path.expanduser(
    "~/Library/CloudStorage/GoogleDrive-changyun1222@gmail.com/"
    "내 드라이브/공시정리/잠정실적_누적.xlsx"
)
STATE_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "earnings_tracker_state.json"
)

HEADERS = [
    "공시일자", "기업명", "업종", "종목코드",
    "당분기 매출액(억원)", "당분기 영업이익(억원)", "당분기 순이익(억원)",
    "매출 QoQ %", "영업이익 QoQ %",
    "전전전분기 매출액(억원)", "전전분기 매출액(억원)", "전분기 매출액(억원)",
    "전전전분기 영업이익(억원)", "전전분기 영업이익(억원)", "전분기 영업이익(억원)",
    "전전전분기 순이익(억원)", "전전분기 순이익(억원)", "전분기 순이익(억원)",
    "기업개요", "DART 링크", "파싱상태",
]

DAILY_HAIKU_LIMIT = 300

CORP_CODE_CACHE_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "dart_corp_codes.json"
)
CORP_CODE_CACHE_TTL_DAYS = 7

INDUSTRY_CACHE_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "dart_industry_cache.json"
)
INDUSTRY_NAME_CACHE_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "dart_industry_name_cache.json"
)
BUSINESS_OVERVIEW_CACHE_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "business_overview_cache.json"
)
BUSINESS_OVERVIEW_TTL_DAYS = 30

REPORT_CODES = {
    "Q1": "11013",
    "H1": "11012",
    "3Q": "11014",
    "ANN": "11011",
}

UNIT_TO_HUNDRED_MILLION_KRW = {
    "조원": 10_000,
    "억원": 1,
    "백만원": 0.01,
    "천원": 0.00001,
    "원": 0.00000001,
}

EARNINGS_PATTERN = re.compile(r"영업\s*\(\s*잠정\s*\)\s*실적")


def _is_earnings_filing(report_nm: str) -> bool:
    if not report_nm:
        return False
    return bool(EARNINGS_PATTERN.search(report_nm))


_anthropic_client = None


def _get_anthropic_client():
    global _anthropic_client
    if _anthropic_client is None:
        _anthropic_client = anthropic.Anthropic(
            api_key=os.environ.get("ANTHROPIC_API_KEY")
        )
    return _anthropic_client


def fetch_filing_body(rcept_no: str) -> str:
    if not DART_API_KEY:
        return ""
    try:
        res = requests.get(
            "https://opendart.fss.or.kr/api/document.xml",
            params={"crtfc_key": DART_API_KEY, "rcept_no": rcept_no},
            timeout=30,
        )
        if res.status_code != 200 or len(res.content) < 100:
            return ""
        zf = zipfile.ZipFile(io.BytesIO(res.content))
        parts = []
        for fname in sorted(zf.namelist()):
            raw = zf.read(fname)
            try:
                text = raw.decode("euc-kr")
            except UnicodeDecodeError:
                text = raw.decode("utf-8", errors="replace")
            parts.append(text)
        return "\n".join(parts)
    except Exception as e:
        print(f"[earnings_tracker] 본문 fetch 실패 ({rcept_no}): {e}")
        return ""


def extract_earnings_tables(html: str) -> str:
    if not html:
        return ""
    soup = BeautifulSoup(html, "html.parser")
    keywords = ["매출", "영업이익", "당기순이익", "순이익", "영업수익"]
    relevant = []
    for table in soup.find_all("table"):
        if any(k in table.get_text() for k in keywords):
            relevant.append(table)
    if not relevant:
        return ""
    parts = []
    for ti, table in enumerate(relevant[:3]):
        rows = []
        for tr in table.find_all("tr"):
            cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
            if any(c for c in cells):
                rows.append(" | ".join(cells))
        parts.append(f"### Table {ti + 1}\n" + "\n".join(rows))
    combined = "\n\n---\n\n".join(parts)
    if len(combined) > 12000:
        combined = combined[:12000]
    return combined


def normalize_to_hundred_million_krw(raw_value, unit):
    if raw_value is None or unit is None:
        return None
    if unit not in UNIT_TO_HUNDRED_MILLION_KRW:
        return None
    multiplier = UNIT_TO_HUNDRED_MILLION_KRW[unit]
    return int(round(raw_value * multiplier))


def extract_financials_with_haiku(table_text: str, corp_name: str) -> dict:
    empty = {"revenue": None, "op_income": None, "net_income": None, "currency_unit": None}
    if not table_text:
        return empty

    prompt = f"""다음은 한국 기업 '{corp_name}'의 잠정실적 공시에 포함된 재무 표입니다.

{table_text}

위 표에서 **당기(=당해실적, 가장 최근 단일 분기) 단독 실적**의 다음 3개 수치와 표의 단위를 JSON으로 반환하라:

[추출 항목]
1. 매출액 (또는 영업수익, 수익) — revenue_raw
2. 영업이익 — op_income_raw
3. 당기순이익 (또는 순이익) — net_income_raw

[중요 규칙 — 반드시 준수]
- "당해실적" 또는 "당기실적" 컬럼만 사용. **"누계실적", "전기실적", "전년동기실적"은 절대 사용 금지**.
- 한국 잠정실적 표는 보통 같은 행에 [당해실적, 전기실적, 전년동기실적, 당기누계, 전년동기누계] 컬럼이 나란히 있음. 그중 가장 첫 컬럼(당해실적)만.
- 숫자는 표에 적힌 그대로. 단위 환산 X. 쉼표 제거. (예: "19,514" → 19514)
- 음수(△ 또는 괄호)는 음수로.
- 단위는 표 위/아래에 명시된 것만 인식: "백만원" / "천원" / "원" / "억원" / "조원" / null.
- 값 못 찾으면 해당 raw 필드 null.
- JSON 외 텍스트 금지. 마크다운 금지.

[자릿수 sanity check — 답변 생성 후 스스로 검증]
- 한국 상장사 분기 매출은 보통 100억~10조 사이. 단위 환산 후 환산값이 이 범위 밖이면 단위 인식이 잘못됐을 가능성. 그런 경우 단위를 다시 봐서 정정.

JSON 형식:
{{"revenue_raw": 19514, "op_income_raw": 2556, "net_income_raw": 1958, "currency_unit": "억원"}}"""

    try:
        client = _get_anthropic_client()
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}],
        )
        text = response.content[0].text.strip()
        text = text.replace("```json", "").replace("```", "").strip()
        brace_start = text.find("{")
        brace_end = text.rfind("}")
        if brace_start != -1 and brace_end != -1:
            text = text[brace_start:brace_end + 1]
        data = json.loads(text)
        unit = data.get("currency_unit")
        return {
            "revenue": normalize_to_hundred_million_krw(data.get("revenue_raw"), unit),
            "op_income": normalize_to_hundred_million_krw(data.get("op_income_raw"), unit),
            "net_income": normalize_to_hundred_million_krw(data.get("net_income_raw"), unit),
            "currency_unit": unit,
        }
    except Exception as e:
        print(f"[earnings_tracker] Haiku 파싱 실패 ({corp_name}): {e}")
        return empty


def _check_and_increment_haiku_counter(state):
    today_str = date.today().isoformat()
    counter_date = state.get("haiku_counter_date")
    counter_value = state.get("haiku_counter_value", 0)
    if counter_date != today_str:
        counter_value = 0
        counter_date = today_str
    if counter_value >= DAILY_HAIKU_LIMIT:
        return False, state
    counter_value += 1
    state["haiku_counter_date"] = counter_date
    state["haiku_counter_value"] = counter_value
    return True, state


def _save_haiku_counter(state):
    full_state = load_state()
    full_state["haiku_counter_date"] = state.get("haiku_counter_date")
    full_state["haiku_counter_value"] = state.get("haiku_counter_value", 0)
    save_state(full_state)


def _fetch_and_cache_corp_codes() -> dict:
    if os.path.exists(CORP_CODE_CACHE_FILE):
        try:
            mtime = datetime.fromtimestamp(os.path.getmtime(CORP_CODE_CACHE_FILE))
            if (datetime.now() - mtime).days < CORP_CODE_CACHE_TTL_DAYS:
                with open(CORP_CODE_CACHE_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            pass
    if not DART_API_KEY:
        return {}
    print("[earnings_tracker] DART corpCode.xml 다운로드 중...")
    res = requests.get(
        "https://opendart.fss.or.kr/api/corpCode.xml",
        params={"crtfc_key": DART_API_KEY},
        timeout=30,
    )
    with zipfile.ZipFile(io.BytesIO(res.content)) as zf:
        with zf.open("CORPCODE.xml") as xf:
            tree = ET.parse(xf)
    mapping = {}
    for elem in tree.getroot().findall("list"):
        stock_code = (elem.findtext("stock_code") or "").strip()
        corp_code = (elem.findtext("corp_code") or "").strip()
        if stock_code and corp_code:
            mapping[stock_code] = corp_code
    with open(CORP_CODE_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(mapping, f)
    print(f"[earnings_tracker] DART corpCode 매핑 {len(mapping)}건 캐시 저장")
    return mapping


def get_corp_code(stock_code: str) -> str | None:
    if not stock_code:
        return None
    mapping = _fetch_and_cache_corp_codes()
    return mapping.get(stock_code.strip())


def get_corp_industry(corp_code: str) -> str | None:
    if not DART_API_KEY or not corp_code:
        return None
    try:
        res = requests.get(
            "https://opendart.fss.or.kr/api/company.json",
            params={"crtfc_key": DART_API_KEY, "corp_code": corp_code},
            timeout=15,
        )
        data = res.json()
    except Exception as e:
        print(f"[earnings_tracker] corpInfo fetch 실패 ({corp_code}): {e}")
        return None
    if data.get("status") != "000":
        return None
    return data.get("induty_code") or None


def get_corp_industry_cached(corp_code: str) -> str | None:
    cache = {}
    if os.path.exists(INDUSTRY_CACHE_FILE):
        try:
            with open(INDUSTRY_CACHE_FILE, "r", encoding="utf-8") as f:
                cache = json.load(f)
        except Exception:
            pass
    if corp_code in cache:
        return cache[corp_code]
    industry = get_corp_industry(corp_code)
    cache[corp_code] = industry
    try:
        with open(INDUSTRY_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False)
    except Exception:
        pass
    return industry


def _translate_industry_code_with_haiku(industry_code: str) -> str | None:
    if not industry_code:
        return None
    prompt = f"""한국표준산업분류(KSIC) 코드 "{industry_code}"에 해당하는 한글 업종명을 반환하라.

규칙:
- 2~5자리 숫자 코드. 예: 264 = 반도체 제조업, 2612 = 평판디스플레이 제조업, 27192 = 측정용기 제조업
- 업종명만 반환. 설명·따옴표·마크다운 금지.
- 코드를 모르면 "기타"라고만 답하라.
- 응답은 JSON 형식: {{"industry_name": "..."}}

답변:"""
    try:
        client = _get_anthropic_client()
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=80,
            messages=[{"role": "user", "content": prompt}],
        )
        text = response.content[0].text.strip()
        text = text.replace("```json", "").replace("```", "").strip()
        data = json.loads(text)
        return data.get("industry_name")
    except Exception as e:
        print(f"[earnings_tracker] 업종명 변환 실패 ({industry_code}): {e}")
        return None


def get_industry_name_cached(industry_code: str, state: dict = None) -> str | None:
    if not industry_code:
        return None
    cache = {}
    if os.path.exists(INDUSTRY_NAME_CACHE_FILE):
        try:
            with open(INDUSTRY_NAME_CACHE_FILE, "r", encoding="utf-8") as f:
                cache = json.load(f)
        except Exception:
            pass
    if industry_code in cache:
        return cache[industry_code]
    if state is not None:
        can_call, _ = _check_and_increment_haiku_counter(state)
        if not can_call:
            return None
    name = _translate_industry_code_with_haiku(industry_code)
    cache[industry_code] = name
    try:
        with open(INDUSTRY_NAME_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
    return name


def fetch_business_section_from_dart(corp_code: str) -> str | None:
    if not corp_code or not DART_API_KEY:
        return None
    end_dt = date.today()
    bgn_dt = end_dt.replace(year=end_dt.year - 2)
    try:
        res = requests.get(
            "https://opendart.fss.or.kr/api/list.json",
            params={
                "crtfc_key": DART_API_KEY,
                "corp_code": corp_code,
                "bgn_de": bgn_dt.strftime("%Y%m%d"),
                "end_de": end_dt.strftime("%Y%m%d"),
                "pblntf_ty": "A",
                "page_count": 20,
            },
            timeout=15,
        )
        data = res.json()
    except Exception as e:
        print(f"[earnings_tracker] 사업보고서 검색 실패 ({corp_code}): {e}")
        return None
    if data.get("status") != "000":
        return None
    items = data.get("list", [])
    rcept_no = None
    for item in items:
        if "사업보고서" in (item.get("report_nm") or ""):
            rcept_no = item.get("rcept_no")
            break
    if not rcept_no:
        return None
    html = fetch_filing_body(rcept_no)
    if not html:
        return None
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text(separator="\n", strip=True)
    start_patterns = [
        r"II\.?\s*사업의\s*내용",
        r"Ⅱ\.?\s*사업의\s*내용",
        r"2\.\s*사업의\s*내용",
    ]
    end_patterns = [
        r"III\.?\s*재무에\s*관한\s*사항",
        r"Ⅲ\.?\s*재무에\s*관한\s*사항",
        r"3\.\s*재무에\s*관한\s*사항",
    ]
    start_idx = None
    for p in start_patterns:
        m = re.search(p, text)
        if m:
            start_idx = m.end()
            break
    if start_idx is None:
        return None
    end_idx = len(text)
    for p in end_patterns:
        m = re.search(p, text[start_idx:])
        if m:
            end_idx = start_idx + m.start()
            break
    section = text[start_idx:end_idx].strip()
    if len(section) > 5000:
        section = section[:5000]
    return section if len(section) > 100 else None


def _summarize_business_with_haiku(corp_name: str, business_section: str) -> str | None:
    if not business_section:
        return None
    prompt = f"""다음은 한국 기업 '{corp_name}'의 사업보고서 '사업의 내용' 섹션입니다.

{business_section}

위 내용을 바탕으로, 이 회사가 무슨 사업을 하는지 **한국어 평문 2~3줄**로 요약하라.

규칙:
- 마크다운(#, *, -, 등) 금지. 평문만.
- 회사 이름은 다시 적지 말것 (이미 컨텍스트에서 알려져 있음).
- 사업 영역, 주요 제품/서비스, 매출 비중이 큰 분야 위주.
- 2~3 문장. 80~150자 정도.

답변:"""
    try:
        client = _get_anthropic_client()
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}],
        )
        text = response.content[0].text.strip()
        text = re.sub(r"^[#*\-\s]+", "", text)
        return text if text else None
    except Exception as e:
        print(f"[earnings_tracker] 기업개요 요약 실패 ({corp_name}): {e}")
        return None


def get_business_overview_cached(stock_code: str, corp_name: str, corp_code: str, state: dict = None) -> str | None:
    if not stock_code:
        return None
    cache = {}
    if os.path.exists(BUSINESS_OVERVIEW_CACHE_FILE):
        try:
            with open(BUSINESS_OVERVIEW_CACHE_FILE, "r", encoding="utf-8") as f:
                cache = json.load(f)
        except Exception:
            pass
    if stock_code in cache:
        try:
            cached_at = date.fromisoformat(cache[stock_code]["cached_at"])
            if (date.today() - cached_at).days < BUSINESS_OVERVIEW_TTL_DAYS:
                return cache[stock_code].get("overview")
        except Exception:
            pass
    if state is not None:
        can_call, _ = _check_and_increment_haiku_counter(state)
        if not can_call:
            return None
    section = fetch_business_section_from_dart(corp_code)
    if not section:
        return None
    overview = _summarize_business_with_haiku(corp_name, section)
    cache[stock_code] = {
        "overview": overview,
        "cached_at": date.today().isoformat(),
    }
    try:
        with open(BUSINESS_OVERVIEW_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
    return overview


def compute_qoq_pct(current, previous) -> float | None:
    if current is None or previous is None:
        return None
    if previous == 0:
        return None
    return round((current - previous) / abs(previous) * 100, 2)


def passes_filter(financials: dict, prev1: dict) -> bool:
    cur_rev = financials.get("revenue")
    prev_rev = prev1.get("revenue")
    cur_op = financials.get("op_income")
    prev_op = prev1.get("op_income")
    rev_up = (cur_rev is not None and prev_rev is not None and cur_rev > prev_rev)
    op_up = (cur_op is not None and prev_op is not None and cur_op > prev_op)
    return rev_up or op_up


def _fetch_fnltt_raw(corp_code: str, year: int, reprt_code: str) -> dict | None:
    for fs_div in ("CFS", "OFS"):
        try:
            res = requests.get(
                "https://opendart.fss.or.kr/api/fnlttSinglAcntAll.json",
                params={
                    "crtfc_key": DART_API_KEY,
                    "corp_code": corp_code,
                    "bsns_year": str(year),
                    "reprt_code": reprt_code,
                    "fs_div": fs_div,
                },
                timeout=15,
            )
            data = res.json()
        except Exception as e:
            print(f"[earnings_tracker] fnltt fetch 실패 ({corp_code} {year} {reprt_code} {fs_div}): {e}")
            return None
        if data.get("status") == "000":
            return data
    return None


def _extract_amounts(data: dict, amount_field: str, sub_field: str | None = None) -> dict:
    """CIS 항목에서 매출/영업이익/순이익 추출. sub_field로 차감도 가능."""
    revenue_keys = ["매출액", "수익(매출액)", "영업수익", "수익"]
    op_income_keys = ["영업이익", "영업이익(손실)", "영업손익"]
    net_income_keys = ["당기순이익", "당기순이익(손실)", "당기순손익",
                       "분기순이익", "분기순이익(손실)", "분기순손익",
                       "반기순이익", "반기순이익(손실)", "반기순손익",
                       "연결당기순이익"]
    result = {"revenue": None, "op_income": None, "net_income": None}
    for item in data.get("list", []):
        if item.get("sj_div") not in ("CIS", "IS"):
            continue
        account_nm = item.get("account_nm", "")
        val_str = (item.get(amount_field) or "").replace(",", "").strip()
        if not val_str or val_str == "-":
            continue
        try:
            val = int(val_str)
        except ValueError:
            continue
        if sub_field:
            sub_str = (item.get(sub_field) or "").replace(",", "").strip()
            if not sub_str or sub_str == "-":
                continue
            try:
                val = val - int(sub_str)
            except ValueError:
                continue
        amount_eokwon = val // 100_000_000
        if result["revenue"] is None and account_nm in revenue_keys:
            result["revenue"] = amount_eokwon
        elif result["op_income"] is None and account_nm in op_income_keys:
            result["op_income"] = amount_eokwon
        elif result["net_income"] is None and account_nm in net_income_keys:
            result["net_income"] = amount_eokwon
    return result


def get_quarter_standalone(stock_code: str, year: int, quarter: str) -> dict:
    """
    특정 기업의 특정 분기 단독 매출/영업이익/순이익 (백만원).
    Q1/Q2/Q3: thstrm_amount가 이미 단독값.
    Q4: ANN.thstrm_amount(연간) - 3Q.thstrm_add_amount(3Q누적) 차감.
    """
    empty = {"revenue": None, "op_income": None, "net_income": None}
    corp_code = get_corp_code(stock_code)
    if not corp_code:
        print(f"[earnings_tracker] corp_code 없음: {stock_code}")
        return empty
    if quarter not in ("Q1", "Q2", "Q3", "Q4"):
        return empty
    if not DART_API_KEY:
        return empty

    if quarter == "Q4":
        ann_data = _fetch_fnltt_raw(corp_code, year, REPORT_CODES["ANN"])
        q3_data = _fetch_fnltt_raw(corp_code, year, REPORT_CODES["3Q"])
        if not ann_data or not q3_data:
            return empty
        ann_vals = _extract_amounts(ann_data, "thstrm_amount")
        q3_cum = _extract_amounts(q3_data, "thstrm_add_amount")
        result = {}
        for k in ("revenue", "op_income", "net_income"):
            a = ann_vals.get(k)
            c = q3_cum.get(k)
            result[k] = (a - c) if (a is not None and c is not None) else None
        return result

    period_map = {"Q1": "Q1", "Q2": "H1", "Q3": "3Q"}
    reprt_code = REPORT_CODES[period_map[quarter]]
    data = _fetch_fnltt_raw(corp_code, year, reprt_code)
    if not data:
        return empty
    return _extract_amounts(data, "thstrm_amount")


def infer_target_quarter(filing_date) -> tuple:
    if isinstance(filing_date, str):
        filing_date = date.fromisoformat(filing_date)
    month = filing_date.month
    year = filing_date.year
    if 1 <= month <= 3:
        return year - 1, "Q4"
    elif 4 <= month <= 6:
        return year, "Q1"
    elif 7 <= month <= 9:
        return year, "Q2"
    else:
        return year, "Q3"


def get_prev_three_quarters(target_year: int, target_quarter: str) -> list:
    order = ["Q1", "Q2", "Q3", "Q4"]
    idx = order.index(target_quarter)
    result = []
    y, q_idx = target_year, idx
    for _ in range(3):
        q_idx -= 1
        if q_idx < 0:
            q_idx = 3
            y -= 1
        result.append((y, order[q_idx]))
    return result


def fetch_historical_quarters(stock_code: str, filing_date) -> dict:
    empty = {"revenue": None, "op_income": None, "net_income": None}
    target_year, target_quarter = infer_target_quarter(filing_date)
    prev_quarters = get_prev_three_quarters(target_year, target_quarter)
    result = {"target_quarter_label": f"{target_year}{target_quarter}"}
    for i, (year, quarter) in enumerate(prev_quarters, start=1):
        try:
            data = get_quarter_standalone(stock_code, year, quarter)
        except Exception as e:
            print(f"[earnings_tracker] {stock_code} {year}{quarter} fetch 에러: {e}")
            data = empty
        result[f"prev{i}"] = data
    return result


def load_state():
    try:
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {"last_run_date": None, "last_run_at": None}


def save_state(state):
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


def fetch_filings(target_date):
    """target_date의 공시 중 잠정실적 공시만 반환."""
    if not DART_API_KEY:
        raise RuntimeError("DART_API_KEY 환경변수가 설정되지 않았습니다.")

    target_str = target_date.strftime("%Y%m%d")
    matched = []
    page_no = 1

    while True:
        try:
            res = requests.get(
                "https://opendart.fss.or.kr/api/list.json",
                params={
                    "crtfc_key": DART_API_KEY,
                    "bgn_de": target_str,
                    "end_de": target_str,
                    "page_count": 100,
                    "page_no": page_no,
                },
                timeout=15,
            )
            data = res.json()
        except Exception as e:
            print(f"[earnings_tracker] DART list.json fetch 실패 (page {page_no}): {e}")
            break

        if data.get("status") != "000":
            if data.get("status") == "013":
                break
            print(f"[earnings_tracker] DART API 에러: {data.get('status')} {data.get('message')}")
            break

        items = data.get("list", [])
        for item in items:
            if _is_earnings_filing(item.get("report_nm", "")):
                matched.append(item)

        total_page = data.get("total_page", 1)
        if page_no >= total_page:
            break
        page_no += 1

    return matched


def append_rows(rows):
    """
    새 행들을 엑셀에 누적 기록.
    같은 (공시일자, 종목코드) 키가 이미 있으면 새 데이터로 덮어쓰기 (정정공시 우선).
    """
    if not rows:
        return {"added": 0, "updated": 0}

    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)

    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        current_headers = [cell.value for cell in ws[1]]
        if len(current_headers) < len(HEADERS):
            for col_idx, header in enumerate(HEADERS, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            print(f"[earnings_tracker] 헤더 마이그레이션: {len(current_headers)}컬럼 → {len(HEADERS)}컬럼")
        existing = {}
        for idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if r and len(r) >= 4 and r[0] and r[3]:
                existing[(str(r[0]), str(r[3]))] = idx
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "잠정실적"
        ws.append(HEADERS)
        existing = {}

    added = 0
    updated = 0
    for row in rows:
        key = (str(row[0]), str(row[3]))
        if key in existing:
            target_row = existing[key]
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=target_row, column=col_idx, value=value)
            updated += 1
        else:
            ws.append(row)
            existing[key] = ws.max_row
            added += 1

    wb.save(EXCEL_PATH)
    return {"added": added, "updated": updated}


def run_daily(force=False, target_date=None):
    """
    하루 1회 실행 진입점.
    target_date=None이면 오늘 기준. 백필 시 특정 날짜 지정 가능 (state 미갱신).
    """
    is_backfill = target_date is not None
    today = target_date or date.today()

    if today.weekday() >= 5:
        msg = f"[earnings_tracker] {today} 주말이므로 skip"
        if is_backfill:
            msg += " (backfill)"
        print(msg)
        return {"status": "skipped_weekend", "date": today.isoformat()}

    if not is_backfill:
        state = load_state()
        if not force and state.get("last_run_date") == today.isoformat():
            print(f"[earnings_tracker] {today} 이미 실행됨 skip")
            return {"status": "skipped_already_run", "date": today.isoformat()}

    mode_label = "[backfill]" if is_backfill else "[daily]"
    print(f"[earnings_tracker] {mode_label} {today} 실행 시작")

    try:
        filings = fetch_filings(today)
        print(f"[earnings_tracker] {mode_label} {today} 공시 {len(filings)}건 발견")

        state = load_state()
        rows = []
        filtered_out = 0
        for f in filings:
            rcept_no = f.get("rcept_no", "")
            corp_name = f.get("corp_name", "")
            stock_code = f.get("stock_code", "")
            corp_code = get_corp_code(stock_code)
            industry_code = get_corp_industry_cached(corp_code) if corp_code else None

            can_call, state = _check_and_increment_haiku_counter(state)
            if not can_call:
                print(f"[earnings_tracker] Haiku 일일 상한({DAILY_HAIKU_LIMIT}) 도달, {corp_name} 파싱 skip")
                financials = {"revenue": None, "op_income": None, "net_income": None, "currency_unit": None}
                parse_status = "limit_reached"
            else:
                html = fetch_filing_body(rcept_no)
                table_text = extract_earnings_tables(html)
                if not table_text:
                    financials = {"revenue": None, "op_income": None, "net_income": None, "currency_unit": None}
                    parse_status = "no_table_found"
                else:
                    financials = extract_financials_with_haiku(table_text, corp_name)
                    if financials["revenue"] is None and financials["op_income"] is None:
                        parse_status = "haiku_failed"
                    else:
                        parse_status = "ok"

            historical = fetch_historical_quarters(stock_code, today)
            prev1 = historical["prev1"]

            if not passes_filter(financials, prev1):
                print(f"[earnings_tracker] 필터 미통과 skip: {corp_name}")
                filtered_out += 1
                continue

            qoq_revenue = compute_qoq_pct(financials.get("revenue"), prev1.get("revenue"))
            qoq_op = compute_qoq_pct(financials.get("op_income"), prev1.get("op_income"))
            industry_name = get_industry_name_cached(industry_code, state)
            overview = get_business_overview_cached(stock_code, corp_name, corp_code, state)

            row = [
                today.isoformat(),
                corp_name,
                industry_name or industry_code,
                stock_code,
                financials["revenue"],
                financials["op_income"],
                financials["net_income"],
                qoq_revenue,
                qoq_op,
                historical["prev3"]["revenue"],
                historical["prev2"]["revenue"],
                historical["prev1"]["revenue"],
                historical["prev3"]["op_income"],
                historical["prev2"]["op_income"],
                historical["prev1"]["op_income"],
                historical["prev3"]["net_income"],
                historical["prev2"]["net_income"],
                historical["prev1"]["net_income"],
                overview,
                f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rcept_no}",
                parse_status,
            ]
            rows.append(row)

        _save_haiku_counter(state)

        result = append_rows(rows)
        print(f"[earnings_tracker] {mode_label} {today} 신규 {result['added']}건, 정정 {result['updated']}건, 필터제외 {filtered_out}건")

        if not is_backfill:
            state = load_state()
            state.update({
                "last_run_date": today.isoformat(),
                "last_run_at": datetime.now().isoformat(),
                "last_run_filings_found": len(filings),
                "last_run_added": result["added"],
                "last_run_updated": result["updated"],
            })
            save_state(state)

        return {
            "status": "ok",
            "date": today.isoformat(),
            "filings_found": len(filings),
            "added": result["added"],
            "updated": result["updated"],
            "mode": "backfill" if is_backfill else "daily",
        }

    except Exception as e:
        import traceback
        print(f"[earnings_tracker] {mode_label} {today} 에러: {e}")
        traceback.print_exc()
        return {"status": "error", "date": today.isoformat(), "message": str(e)}


def run_backfill(days):
    """오늘 포함 N일치 백필. 주말 자동 skip."""
    if days < 1:
        return []

    today = date.today()
    results = []

    print(f"[earnings_tracker] 백필 시작: {days}일")

    for i in range(days):
        target = today - timedelta(days=i)
        result = run_daily(target_date=target)
        results.append(result)

    total_added = sum(r.get("added", 0) for r in results if r.get("status") == "ok")
    total_updated = sum(r.get("updated", 0) for r in results if r.get("status") == "ok")
    print(f"[earnings_tracker] 백필 완료: 총 신규 {total_added}건, 정정 {total_updated}건")

    return results


if __name__ == "__main__":
    import sys

    if "--verify-q-semantics" in sys.argv:
        quarters = {}
        for q in ["Q1", "Q2", "Q3", "Q4"]:
            quarters[q] = get_quarter_standalone("006400", 2025, q)
        corp_code = get_corp_code("006400")
        ann_data = _fetch_fnltt_raw(corp_code, 2025, REPORT_CODES["ANN"])
        ann_amounts = _extract_amounts(ann_data, "thstrm_amount") if ann_data else None
        total = {"revenue": 0, "op_income": 0, "net_income": 0}
        valid = True
        for q in ["Q1", "Q2", "Q3", "Q4"]:
            for k in total:
                v = quarters[q].get(k)
                if v is None:
                    valid = False
                else:
                    total[k] += v
        print(f"\n=== 삼성SDI 006400 2025 분기 검증 ===")
        print(f"{'분기':<10} {'매출액':>15} {'영업이익':>15} {'순이익':>15}")
        for q in ["Q1", "Q2", "Q3", "Q4"]:
            d = quarters[q]
            print(f"{q:<10} {str(d.get('revenue')):>15} {str(d.get('op_income')):>15} {str(d.get('net_income')):>15}")
        if valid:
            print(f"{'합계':<10} {total['revenue']:>15} {total['op_income']:>15} {total['net_income']:>15}")
        else:
            print(f"{'합계':<10} {'(일부 None)':>15}")
        a = ann_amounts or {}
        print(f"{'ANN':<10} {str(a.get('revenue', '?')):>15} {str(a.get('op_income', '?')):>15} {str(a.get('net_income', '?')):>15}")
        sys.exit(0)

    for arg in sys.argv[1:]:
        if arg.startswith("--test-quarter="):
            parts = arg.split("=", 1)[1].split(",")
            if len(parts) != 3:
                print("형식: --test-quarter=종목코드,연도,분기 (예: --test-quarter=047040,2025,Q4)")
                sys.exit(1)
            stock_code, year, quarter = parts
            result = get_quarter_standalone(stock_code, int(year), quarter)
            print(f"[test] {stock_code} {year} {quarter}: {result}")
            sys.exit(0)

    backfill_days = None
    for arg in sys.argv[1:]:
        if arg.startswith("--backfill-days="):
            try:
                backfill_days = int(arg.split("=", 1)[1])
            except ValueError:
                print(f"잘못된 값: {arg}. --backfill-days=숫자 형식 사용")
                sys.exit(1)

    if backfill_days is not None:
        results = run_backfill(backfill_days)
        print(f"[earnings_tracker] 백필 결과 (일자별):")
        for r in results:
            print(f"  {r.get('date')}: {r.get('status')} added={r.get('added', 0)} updated={r.get('updated', 0)}")
    else:
        force = "--force" in sys.argv
        result = run_daily(force=force)
        print(f"[earnings_tracker] 결과: {result}")
